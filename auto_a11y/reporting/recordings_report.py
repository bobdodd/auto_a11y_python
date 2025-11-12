"""
Recordings report generator for AutoA11y

Generates reports from manual accessibility audit recordings and lived experience testing.
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
from collections import defaultdict

logger = logging.getLogger(__name__)


class RecordingsReportGenerator:
    """Generate reports from accessibility recordings"""

    def __init__(self, db, config: dict):
        """
        Initialize report generator

        Args:
            db: Database connection
            config: Configuration dictionary
        """
        self.db = db
        self.config = config
        self.reports_dir = Path(config.get('REPORTS_DIR', 'reports'))
        self.reports_dir.mkdir(exist_ok=True)

    def generate_project_recordings_report(
        self,
        project_id: str,
        format: str = 'html',
        include_summary: bool = True,
        include_timecodes: bool = True,
        include_wcag: bool = True,
        group_by_touchpoint: bool = True
    ) -> str:
        """
        Generate a report for all recordings in a project

        Args:
            project_id: Project ID
            format: Output format (html, pdf, xlsx)
            include_summary: Include executive summary
            include_timecodes: Include timecodes for issues
            include_wcag: Include WCAG criteria mappings
            group_by_touchpoint: Group issues by touchpoint

        Returns:
            Path to generated report file
        """
        # Get project
        project = self.db.get_project(project_id)
        if not project:
            raise ValueError(f"Project not found: {project_id}")

        # Get all recordings for this project
        recordings = self.db.get_recordings(project_id=project_id)
        if not recordings:
            raise ValueError(f"No recordings found for project: {project.name}")

        # Collect all issues from recordings
        all_issues = []
        issues_by_recording = {}

        for recording in recordings:
            issues = self.db.get_recording_issues_for_recording(recording.recording_id)
            issues_by_recording[recording.recording_id] = issues
            all_issues.extend(issues)

        # Calculate summary statistics
        stats = self._calculate_statistics(recordings, all_issues)

        # Group issues if requested
        issues_by_touchpoint = None
        if group_by_touchpoint:
            issues_by_touchpoint = self._group_by_touchpoint(all_issues)

        # Generate report based on format
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        project_name_safe = project.name.replace(' ', '_').replace('/', '_')

        if format == 'html':
            filename = f"{project_name_safe}_recordings_{timestamp}.html"
            report_path = self.reports_dir / filename
            self._generate_html_report(
                report_path,
                project,
                recordings,
                all_issues,
                issues_by_touchpoint,
                stats,
                include_summary,
                include_timecodes,
                include_wcag
            )
        elif format == 'pdf':
            filename = f"{project_name_safe}_recordings_{timestamp}.pdf"
            report_path = self.reports_dir / filename
            self._generate_pdf_report(
                report_path,
                project,
                recordings,
                all_issues,
                issues_by_touchpoint,
                stats,
                include_summary,
                include_timecodes,
                include_wcag
            )
        elif format == 'xlsx':
            filename = f"{project_name_safe}_recordings_{timestamp}.xlsx"
            report_path = self.reports_dir / filename
            self._generate_xlsx_report(
                report_path,
                project,
                recordings,
                all_issues,
                issues_by_touchpoint,
                stats,
                include_timecodes,
                include_wcag
            )
        else:
            raise ValueError(f"Unsupported format: {format}")

        logger.info(f"Generated recordings report: {report_path}")
        return str(report_path)

    def _calculate_statistics(self, recordings, issues):
        """Calculate summary statistics"""
        total_recordings = len(recordings)
        total_issues = len(issues)

        # Count by impact
        high_count = sum(1 for issue in issues if issue.impact.value.lower() in ['high', 'critical'])
        medium_count = sum(1 for issue in issues if issue.impact.value.lower() in ['medium', 'moderate'])
        low_count = sum(1 for issue in issues if issue.impact.value.lower() == 'low')

        # Count by recording type
        audit_recordings = sum(1 for r in recordings if r.recording_type.value == 'audit')
        lived_exp_recordings = len(recordings) - audit_recordings

        # Count touchpoints
        touchpoints = set(issue.touchpoint for issue in issues if issue.touchpoint)

        return {
            'total_recordings': total_recordings,
            'total_issues': total_issues,
            'high_count': high_count,
            'medium_count': medium_count,
            'low_count': low_count,
            'audit_recordings': audit_recordings,
            'lived_experience_recordings': lived_exp_recordings,
            'unique_touchpoints': len(touchpoints),
            'touchpoints': sorted(touchpoints) if touchpoints else []
        }

    def _group_by_touchpoint(self, issues):
        """Group issues by touchpoint"""
        grouped = defaultdict(list)
        for issue in issues:
            touchpoint = issue.touchpoint or "General"
            grouped[touchpoint].append(issue)
        return dict(sorted(grouped.items()))

    def _generate_html_report(
        self,
        output_path,
        project,
        recordings,
        issues,
        issues_by_touchpoint,
        stats,
        include_summary,
        include_timecodes,
        include_wcag
    ):
        """Generate HTML report with embedded Bootstrap for offline use"""
        from pathlib import Path
        import urllib.request

        # Get path to static files
        static_dir = Path(__file__).parent.parent / 'web' / 'static'

        # Read Bootstrap CSS - use CDN as fallback if local file doesn't exist
        bootstrap_css = ''
        bootstrap_css_path = static_dir / 'css' / 'bootstrap.min.css'
        if bootstrap_css_path.exists():
            bootstrap_css = bootstrap_css_path.read_text(encoding='utf-8')
        else:
            # Fetch from CDN as fallback
            try:
                with urllib.request.urlopen('https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css') as response:
                    bootstrap_css = response.read().decode('utf-8')
            except Exception as e:
                logger.warning(f"Could not fetch Bootstrap CSS: {e}")

        # For Bootstrap Icons, we need to either:
        # 1. Remove icon CSS entirely and not use icons, or
        # 2. Use a CDN version that works offline (not possible), or
        # 3. Convert fonts to base64 data URIs (complex)
        # For now, we'll skip Bootstrap Icons to avoid font loading errors
        # The report will work fine without icons
        bootstrap_icons_css = ''

        # Alternative: If you have font files, convert them to base64 and embed
        # For now, just use fallback unicode symbols instead of icon fonts

        # Read Bootstrap JS
        bootstrap_js = ''
        bootstrap_js_path = static_dir / 'js' / 'bootstrap.bundle.min.js'
        if bootstrap_js_path.exists():
            bootstrap_js = bootstrap_js_path.read_text(encoding='utf-8')
        else:
            # Fetch from CDN as fallback
            try:
                with urllib.request.urlopen('https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/js/bootstrap.bundle.min.js') as response:
                    bootstrap_js = response.read().decode('utf-8')
            except Exception as e:
                logger.warning(f"Could not fetch Bootstrap JS: {e}")

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Recordings Report - {project.name}</title>

    <!-- Inlined Bootstrap CSS -->
    <style>
{bootstrap_css}
    </style>

    <!-- Inlined Bootstrap Icons CSS -->
    <style>
{bootstrap_icons_css}
    </style>

    <!-- Custom styles -->
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 40px;
            line-height: 1.6;
        }}
        h1 {{
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }}
        h2 {{
            color: #34495e;
            margin-top: 30px;
            border-bottom: 2px solid #95a5a6;
            padding-bottom: 5px;
        }}
        h3 {{
            color: #7f8c8d;
        }}
        .summary {{
            background: #ecf0f1;
            padding: 20px;
            border-radius: 5px;
            margin: 20px 0;
        }}
        .stat {{
            display: inline-block;
            margin: 10px 20px 10px 0;
        }}
        .stat-value {{
            font-size: 24px;
            font-weight: bold;
            color: #2980b9;
        }}
        .stat-label {{
            color: #7f8c8d;
            font-size: 14px;
        }}
        .recording {{
            background: #fff;
            border: 1px solid #ddd;
            border-radius: 5px;
            padding: 15px;
            margin: 15px 0;
        }}
        .touchpoint-section {{
            margin-bottom: 30px;
        }}
        .touchpoint-header {{
            background: #f8f9fa;
            padding: 10px 15px;
            border-radius: 5px;
            margin-bottom: 15px;
        }}
        .recording-reference {{
            background: #e7f3ff;
            padding: 10px;
            border-radius: 5px;
            margin-bottom: 15px;
            border-left: 3px solid #3498db;
        }}
        .accordion-item {{
            margin-bottom: 10px;
        }}
        .accordion-button:not(.collapsed) {{
            background-color: #e7f5ff;
            color: #0c63e4;
        }}
        .footer {{
            margin-top: 50px;
            padding-top: 20px;
            border-top: 1px solid #ddd;
            color: #7f8c8d;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container-fluid">
        <h1>Recordings Report: {project.name}</h1>
        <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
"""

        if include_summary:
            html += f"""
    <div class="summary">
        <h2>Executive Summary</h2>
        <div class="stat">
            <div class="stat-value">{stats['total_recordings']}</div>
            <div class="stat-label">Total Recordings</div>
        </div>
        <div class="stat">
            <div class="stat-value">{stats['total_issues']}</div>
            <div class="stat-label">Total Issues</div>
        </div>
        <div class="stat">
            <div class="stat-value">{stats['high_count']}</div>
            <div class="stat-label">High Impact</div>
        </div>
        <div class="stat">
            <div class="stat-value">{stats['medium_count']}</div>
            <div class="stat-label">Medium Impact</div>
        </div>
        <div class="stat">
            <div class="stat-value">{stats['low_count']}</div>
            <div class="stat-label">Low Impact</div>
        </div>
        <div class="stat">
            <div class="stat-value">{stats['audit_recordings']}</div>
            <div class="stat-label">Audit Recordings</div>
        </div>
        <div class="stat">
            <div class="stat-value">{stats['lived_experience_recordings']}</div>
            <div class="stat-label">Lived Experience</div>
        </div>
        <div class="stat">
            <div class="stat-value">{stats['unique_touchpoints']}</div>
            <div class="stat-label">Touchpoints</div>
        </div>
    </div>
"""

        # Add issues grouped by touchpoint with nested accordions
        if issues_by_touchpoint:
            html += "\n    <h2>Issues by Touchpoint</h2>\n"

            # Create a mapping of recording IDs to recording objects for quick lookup
            recordings_map = {r.recording_id: r for r in recordings}

            # Outer accordion for touchpoints
            html += """
    <div class="accordion" id="touchpoint-accordion">
"""

            for tp_idx, (touchpoint, touchpoint_issues) in enumerate(issues_by_touchpoint.items(), 1):
                touchpoint_id = touchpoint.replace(' ', '-').replace('/', '-').replace('(', '').replace(')', '').replace('.', '')

                html += f"""
        <div class="accordion-item">
            <h2 class="accordion-header">
                <button class="accordion-button collapsed" type="button" data-bs-toggle="collapse" data-bs-target="#touchpoint-{touchpoint_id}" aria-expanded="false" aria-controls="touchpoint-{touchpoint_id}">
                    📁 {touchpoint}
                    <span class="badge bg-secondary ms-2">{len(touchpoint_issues)} issues</span>
                </button>
            </h2>
            <div id="touchpoint-{touchpoint_id}" class="accordion-collapse collapse" data-bs-parent="#touchpoint-accordion">
                <div class="accordion-body">
                    <!-- Inner accordion for issues within this touchpoint -->
                    <div class="accordion" id="accordion-{touchpoint_id}">
"""

                for idx, issue in enumerate(touchpoint_issues, 1):
                    issue_id = f"issue-{idx}-{touchpoint_id}"

                    # Get the recording this issue is from
                    recording = recordings_map.get(issue.recording_id)
                    recording_title = recording.title if recording else "Unknown Recording"
                    recording_type = recording.recording_type.value.replace('_', ' ').title() if recording else "Unknown"

                    # Map impact to Bootstrap badge colors
                    impact_badge_map = {
                        'high': 'danger',
                        'critical': 'danger',
                        'medium': 'warning',
                        'moderate': 'warning',
                        'low': 'secondary'
                    }
                    badge_color = impact_badge_map.get(issue.impact.value.lower(), 'secondary')

                    html += f"""
                        <div class="accordion-item">
                            <h2 class="accordion-header">
                                <button class="accordion-button collapsed" type="button" data-bs-toggle="collapse" data-bs-target="#{issue_id}" aria-expanded="false" aria-controls="{issue_id}">
                                    <span class="badge bg-{badge_color} me-2">{issue.impact.value.upper()}</span>
                                    <strong>{issue.title}</strong>
                                    <small class="text-muted ms-2">({recording_title} - {issue.recording_id})</small>
                                </button>
                            </h2>
                            <div id="{issue_id}" class="accordion-collapse collapse" data-bs-parent="#accordion-{touchpoint_id}">
                                <div class="accordion-body">
                                    <div class="recording-reference">
                                        <strong>Recording:</strong> {recording_title}<br>
                                        <strong>Type:</strong> {recording_type}<br>
                                        <strong>Recording ID:</strong> <code>{issue.recording_id}</code>
                                    </div>
"""

                    if issue.what:
                        html += f"                                    <p><strong>What:</strong> {issue.what}</p>\n"
                    if issue.why:
                        html += f"                                    <p><strong>Why:</strong> {issue.why}</p>\n"
                    if issue.who:
                        html += f"                                    <p><strong>Who is affected:</strong> {issue.who}</p>\n"
                    if issue.remediation:
                        html += f"                                    <p><strong>How to fix:</strong> {issue.remediation}</p>\n"

                    if include_wcag and issue.wcag:
                        html += "                                    <p><strong>WCAG:</strong> "
                        for wcag_ref in issue.wcag:
                            html += f'<span class="badge bg-primary me-1">{wcag_ref.criteria} (Level {wcag_ref.level})</span> '
                        html += "</p>\n"

                    if include_timecodes and issue.timecodes:
                        html += "                                    <p><strong>Timecodes:</strong> "
                        for tc in issue.timecodes:
                            html += f'<span class="badge bg-secondary me-1">{tc.start} - {tc.end}</span> '
                        html += "</p>\n"

                    html += """
                                </div>
                            </div>
                        </div>
"""

                html += """
                    </div>
                </div>
            </div>
        </div>
"""

            html += """
    </div>
"""

        # Add recordings list
        html += "\n    <h2>Recordings</h2>\n"
        for recording in recordings:
            recording_type_badge = 'audit' if recording.recording_type.value == 'audit' else 'lived-exp'
            html += f"""
    <div class="recording">
        <h3>{recording.title}</h3>
        <p>
            <span class="badge badge-{recording_type_badge}">{recording.recording_type.value.replace('_', ' ').title()}</span>
            <strong>ID:</strong> {recording.recording_id}<br>
            <strong>Auditor/Tester:</strong> {recording.auditor_name or 'N/A'}<br>
            <strong>Date:</strong> {recording.recorded_date.strftime('%Y-%m-%d') if recording.recorded_date else 'N/A'}<br>
            <strong>Issues:</strong> {recording.total_issues} total
            ({recording.high_impact_count} high, {recording.medium_impact_count} medium, {recording.low_impact_count} low)
        </p>
    </div>
"""

        html += f"""
    <div class="footer">
        <p>Auto A11y © 2025 | Automated Accessibility Testing</p>
    </div>
    </div>

    <!-- Inlined Bootstrap JS Bundle -->
    <script>
{bootstrap_js}
    </script>
</body>
</html>
"""

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)

    def _generate_pdf_report(self, output_path, project, recordings, issues, issues_by_touchpoint, stats, include_summary, include_timecodes, include_wcag):
        """Generate PDF report (placeholder - would use library like ReportLab or WeasyPrint)"""
        # For now, generate HTML and note that PDF conversion would be done here
        html_path = str(output_path).replace('.pdf', '.html')
        self._generate_html_report(
            html_path,
            project,
            recordings,
            issues,
            issues_by_touchpoint,
            stats,
            include_summary,
            include_timecodes,
            include_wcag
        )
        # TODO: Convert HTML to PDF using WeasyPrint or similar
        logger.warning("PDF generation not fully implemented - HTML report generated instead")
        import shutil
        shutil.copy(html_path, output_path.replace('.pdf', '_temp.html'))

    def _generate_xlsx_report(self, output_path, project, recordings, issues, issues_by_touchpoint, stats, include_timecodes, include_wcag):
        """Generate Excel report (placeholder - would use library like openpyxl)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()

            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            ws_summary['A1'] = "Recordings Report"
            ws_summary['A1'].font = Font(size=16, bold=True)
            ws_summary['A3'] = "Project:"
            ws_summary['B3'] = project.name
            ws_summary['A4'] = "Generated:"
            ws_summary['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            ws_summary['A6'] = "Total Recordings:"
            ws_summary['B6'] = stats['total_recordings']
            ws_summary['A7'] = "Total Issues:"
            ws_summary['B7'] = stats['total_issues']
            ws_summary['A8'] = "High Impact:"
            ws_summary['B8'] = stats['high_count']
            ws_summary['A9'] = "Medium Impact:"
            ws_summary['B9'] = stats['medium_count']
            ws_summary['A10'] = "Low Impact:"
            ws_summary['B10'] = stats['low_count']

            # Issues sheet
            ws_issues = wb.create_sheet("Issues")
            headers = ['Title', 'Impact', 'Touchpoint', 'What', 'Why', 'Who', 'Remediation']
            if include_wcag:
                headers.append('WCAG')
            if include_timecodes:
                headers.append('Timecodes')

            for col, header in enumerate(headers, 1):
                cell = ws_issues.cell(1, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

            for row, issue in enumerate(issues, 2):
                ws_issues.cell(row, 1, issue.title)
                ws_issues.cell(row, 2, issue.impact.value.upper())
                ws_issues.cell(row, 3, issue.touchpoint or "General")
                ws_issues.cell(row, 4, issue.what or "")
                ws_issues.cell(row, 5, issue.why or "")
                ws_issues.cell(row, 6, issue.who or "")
                ws_issues.cell(row, 7, issue.remediation or "")

                col = 8
                if include_wcag:
                    wcag_str = ", ".join([f"{w.criteria} (Level {w.level})" for w in issue.wcag]) if issue.wcag else ""
                    ws_issues.cell(row, col, wcag_str)
                    col += 1

                if include_timecodes:
                    tc_str = ", ".join([f"{tc.start}-{tc.end}" for tc in issue.timecodes]) if issue.timecodes else ""
                    ws_issues.cell(row, col, tc_str)

            # Recordings sheet
            ws_recordings = wb.create_sheet("Recordings")
            rec_headers = ['ID', 'Title', 'Type', 'Auditor/Tester', 'Date', 'Total Issues', 'High', 'Medium', 'Low']
            for col, header in enumerate(rec_headers, 1):
                cell = ws_recordings.cell(1, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

            for row, recording in enumerate(recordings, 2):
                ws_recordings.cell(row, 1, recording.recording_id)
                ws_recordings.cell(row, 2, recording.title)
                ws_recordings.cell(row, 3, recording.recording_type.value.replace('_', ' ').title())
                ws_recordings.cell(row, 4, recording.auditor_name or "N/A")
                ws_recordings.cell(row, 5, recording.recorded_date.strftime('%Y-%m-%d') if recording.recorded_date else "N/A")
                ws_recordings.cell(row, 6, recording.total_issues)
                ws_recordings.cell(row, 7, recording.high_impact_count)
                ws_recordings.cell(row, 8, recording.medium_impact_count)
                ws_recordings.cell(row, 9, recording.low_impact_count)

            wb.save(output_path)

        except ImportError:
            logger.error("openpyxl not installed - cannot generate Excel report")
            raise
