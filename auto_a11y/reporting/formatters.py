"""
Report formatters for different output formats
"""

import json
import csv
from typing import Dict, Any, List
from datetime import datetime
from pathlib import Path
import logging
from auto_a11y.reporting.comprehensive_report import ComprehensiveReportGenerator
from auto_a11y.reporting.issue_catalog import IssueCatalog
from io import StringIO

logger = logging.getLogger(__name__)


class BaseFormatter:
    """Base class for report formatters"""
    
    def __init__(self, config: Dict[str, Any]):
        """Initialize formatter with config"""
        self.config = config
        self.extension = 'txt'
    
    def format_page_report(self, data: Dict[str, Any]) -> str:
        """Format page report data"""
        raise NotImplementedError
    
    def format_website_report(self, data: Dict[str, Any]) -> str:
        """Format website report data"""
        raise NotImplementedError
    
    def format_project_report(self, data: Dict[str, Any]) -> str:
        """Format project report data"""
        raise NotImplementedError
    
    def format_summary_report(self, data: Dict[str, Any]) -> str:
        """Format summary report data"""
        raise NotImplementedError


class HTMLFormatter(BaseFormatter):
    """HTML report formatter"""
    
    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.extension = 'html'
        # Pass Claude API key if available
        claude_api_key = config.get('CLAUDE_API_KEY')
        self.comprehensive_generator = ComprehensiveReportGenerator(claude_api_key=claude_api_key)
    
    def format_all_projects_report(self, data: Dict[str, Any]) -> str:
        """Format report for all projects as HTML"""
        
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{data.get('title', 'All Projects Accessibility Report')}</title>
    {self._get_css()}
</head>
<body>
    <div class="container">
        <header>
            <h1>{data.get('title', 'All Projects Accessibility Report')}</h1>
            <div class="metadata">
                <p><strong>Total Projects:</strong> {data['summary']['total_projects']}</p>
                <p><strong>Generated:</strong> {data['generated_at']}</p>
            </div>
        </header>
        
        <section class="summary">
            <h2>Overall Summary</h2>
            <div class="stats-grid">
                <div class="stat-card">
                    <h3>{data['summary']['total_projects']}</h3>
                    <p>Projects</p>
                </div>
                <div class="stat-card">
                    <h3>{data['summary']['total_websites']}</h3>
                    <p>Websites</p>
                </div>
                <div class="stat-card">
                    <h3>{data['summary']['total_pages']}</h3>
                    <p>Total Pages</p>
                </div>
                <div class="stat-card">
                    <h3>{data['summary']['total_tested']}</h3>
                    <p>Pages Tested</p>
                </div>
                <div class="stat-card violation">
                    <h3>{data['summary']['total_violations']}</h3>
                    <p>Total Violations</p>
                </div>
                <div class="stat-card warning">
                    <h3>{data['summary']['total_warnings']}</h3>
                    <p>Total Warnings</p>
                </div>
            </div>
        </section>
        
        <section class="projects">
            <h2>Projects Breakdown</h2>
            {"".join(self._format_project_section(p) for p in data['projects'])}
        </section>
        
        {self._get_footer()}
    </div>
</body>
</html>"""
        
        return html
    
    def _format_project_section(self, project_data: Dict[str, Any]) -> str:
        """Format a single project section for all projects report"""
        project = project_data['project']
        stats = project_data['stats']
        
        websites_html = ""
        for website_data in project_data['websites']:
            website = website_data['website']
            websites_html += f"""
            <div class="website-item">
                <h4>{website.get('name', website.get('url', 'Unknown'))}</h4>
                <p>Pages: {website_data['pages']} | Tested: {website_data['tested']} | 
                   Violations: {website_data['violations']} | Warnings: {website_data['warnings']}</p>
            </div>
            """
        
        return f"""
        <div class="project-section">
            <h3>{project.get('name', 'Unnamed Project')}</h3>
            <p>{project.get('description', '')}</p>
            <div class="project-stats">
                <span>Websites: {stats['website_count']}</span>
                <span>Pages: {stats['total_pages']}</span>
                <span>Tested: {stats['tested_pages']}</span>
                <span>Coverage: {stats['test_coverage']:.1f}%</span>
            </div>
            <div class="websites-list">
                {websites_html if websites_html else '<p>No websites in this project</p>'}
            </div>
        </div>
        """
    
    def format_page_report(self, data: Dict[str, Any]) -> str:
        """Generate HTML report for a page"""

        # Check for page state information
        test_result = data.get('test_result', {})
        page_state_info = ""
        if test_result and test_result.get('session_id'):
            page_state = test_result.get('page_state', {})
            if isinstance(page_state, dict):
                state_desc = page_state.get('description', '')
            else:
                state_desc = ''

            if state_desc:
                page_state_info = f"""
                <p><strong>Page State:</strong> {state_desc}</p>
                <p><strong>Session ID:</strong> {test_result.get('session_id', '')}</p>
                """

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Accessibility Report - {data['page']['url']}</title>
    {self._get_css()}
</head>
<body>
    <div class="container">
        <header>
            <h1>Accessibility Report</h1>
            <div class="metadata">
                <p><strong>Page:</strong> <a href="{data['page']['url']}" target="_blank">{data['page']['url']}</a></p>
                <p><strong>Website:</strong> {data['website']['name']}</p>
                <p><strong>Project:</strong> {data['project']['name']}</p>
                <p><strong>Generated:</strong> {data['generated_at']}</p>
                {page_state_info}
            </div>
        </header>

        {self._format_summary_section(data['statistics'])}

        {self._format_violations_section(data.get('violations', []))}

        {self._format_warnings_section(data.get('warnings', []))}

        {self._format_info_section(data.get('info', []))}

        {self._format_discovery_section(data.get('discovery', []))}

        {self._format_ai_findings_section(data.get('ai_findings', []))}

        {self._format_passes_section(data.get('passes', []))}

        <footer>
            <p>Generated by Auto A11y Python - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </footer>
    </div>
</body>
</html>"""

        return html
    
    def format_website_report(self, data: Dict[str, Any]) -> str:
        """Generate HTML report for a website"""
        
        # Use comprehensive report generator for website reports
        return self.comprehensive_generator.generate_comprehensive_html(data)
    
    def format_project_report(self, data: Dict[str, Any]) -> str:
        """Generate HTML report for a project"""
        
        # Use comprehensive report generator for project reports
        return self.comprehensive_generator.generate_comprehensive_html(data)
    
    def format_summary_report(self, data: Dict[str, Any]) -> str:
        """Generate executive summary report"""
        
        projects_html = ""
        for project in data['projects']:
            projects_html += f"""
            <tr>
                <td>{project['name']}</td>
                <td>{project['websites']}</td>
                <td>{project['pages_tested']}</td>
                <td class="violations">{project['violations']}</td>
            </tr>"""
        
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Accessibility Executive Summary</title>
    {self._get_css()}
</head>
<body>
    <div class="container">
        <header>
            <h1>Accessibility Executive Summary</h1>
            <p class="generated">Generated: {data['generated_at']}</p>
        </header>
        
        <section class="overview">
            <h2>Overview</h2>
            <div class="stats-grid">
                <div class="stat-card">
                    <h3>{len(data['projects'])}</h3>
                    <p>Projects</p>
                </div>
                <div class="stat-card">
                    <h3>{data['total_pages_tested']}</h3>
                    <p>Pages Tested</p>
                </div>
                <div class="stat-card violations">
                    <h3>{data['total_violations']}</h3>
                    <p>Total Violations</p>
                </div>
            </div>
        </section>
        
        <section class="projects">
            <h2>Projects</h2>
            <table>
                <thead>
                    <tr>
                        <th>Project</th>
                        <th>Websites</th>
                        <th>Pages Tested</th>
                        <th>Violations</th>
                    </tr>
                </thead>
                <tbody>
                    {projects_html}
                </tbody>
            </table>
        </section>
        
        <footer>
            <p>Generated by Auto A11y Python - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </footer>
    </div>
</body>
</html>"""
        
        return html
    
    def _get_footer(self) -> str:
        """Get footer for HTML reports"""
        return """
        <footer>
            <p>Generated by Auto A11y - Accessibility Testing Tool</p>
            <p>Report generated on """ + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + """</p>
        </footer>
        """
    
    def _get_css(self) -> str:
        """Get CSS styles for HTML reports"""
        return """
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            background: #f5f5f5;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            background: white;
            box-shadow: 0 0 10px rgba(0,0,0,0.1);
        }
        header {
            border-bottom: 2px solid #007bff;
            padding-bottom: 20px;
            margin-bottom: 30px;
        }
        h1 { color: #007bff; margin-bottom: 10px; }
        h2 { color: #333; margin: 20px 0 15px; border-bottom: 1px solid #e0e0e0; padding-bottom: 5px; }
        h3 { color: #555; margin: 15px 0 10px; }
        
        .metadata {
            background: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            margin: 15px 0;
        }
        .metadata p { margin: 5px 0; }
        
        .stats-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }
        
        .stat-card {
            background: #f8f9fa;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
            border: 1px solid #dee2e6;
        }
        .stat-card h3 {
            font-size: 2em;
            color: #007bff;
            margin: 0;
        }
        .stat-card.violations h3 { color: #dc3545; }
        .stat-card.warnings h3 { color: #ffc107; }
        .stat-card.passes h3 { color: #28a745; }
        
        .violation, .warning, .pass, .ai-finding {
            background: #fff;
            border-left: 4px solid #dc3545;
            padding: 15px;
            margin: 15px 0;
            border-radius: 4px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .warning { border-left-color: #ffc107; }
        .pass { border-left-color: #28a745; }
        .ai-finding { border-left-color: #6f42c1; }
        
        .violation h4, .warning h4, .pass h4, .ai-finding h4 {
            margin-bottom: 10px;
            color: #333;
        }
        
        .impact {
            display: inline-block;
            padding: 2px 8px;
            border-radius: 3px;
            font-size: 0.85em;
            font-weight: bold;
            margin-left: 10px;
        }
        .impact.critical { background: #dc3545; color: white; }
        .impact.serious { background: #fd7e14; color: white; }
        .impact.moderate { background: #ffc107; color: #333; }
        .impact.minor { background: #6c757d; color: white; }
        
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }
        th, td {
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #dee2e6;
        }
        th {
            background: #f8f9fa;
            font-weight: bold;
            color: #495057;
        }
        tr:hover { background: #f8f9fa; }
        
        .code {
            background: #f4f4f4;
            padding: 10px;
            border-radius: 4px;
            font-family: 'Courier New', monospace;
            font-size: 0.9em;
            overflow-x: auto;
            margin: 10px 0;
        }
        
        footer {
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #dee2e6;
            text-align: center;
            color: #6c757d;
        }
        
        a { color: #007bff; text-decoration: none; }
        a:hover { text-decoration: underline; }
    </style>
        """
    
    def _format_summary_section(self, stats: Dict) -> str:
        """Format summary statistics section"""
        return f"""
        <section class="summary">
            <h2>Summary</h2>
            <div class="stats-grid">
                <div class="stat-card violations">
                    <h3>{stats['violations']}</h3>
                    <p>Violations</p>
                </div>
                <div class="stat-card warnings">
                    <h3>{stats['warnings']}</h3>
                    <p>Warnings</p>
                </div>
                <div class="stat-card passes">
                    <h3>{stats['passes']}</h3>
                    <p>Passes</p>
                </div>
                <div class="stat-card">
                    <h3>{stats['duration_ms']}ms</h3>
                    <p>Test Duration</p>
                </div>
            </div>
        </section>"""
    
    def _format_violations_section(self, violations: List[Dict]) -> str:
        """Format violations section"""
        if not violations:
            return ""

        html = "<section class='violations'><h2>Violations</h2>"
        for v in violations:
            impact_class = v.get('impact', 'moderate').lower()
            metadata = v.get('metadata', {})

            # Build metadata display
            metadata_html = ""
            if metadata.get('breakpoint'):
                metadata_html += f"<p><strong>Breakpoint:</strong> {metadata.get('breakpoint')}px width</p>"
            if metadata.get('pseudoclass'):
                metadata_html += f"<p><strong>CSS State:</strong> {metadata.get('pseudoclass')}</p>"

            html += f"""
            <div class="violation">
                <h4>{v.get('rule_id', 'Unknown')}
                    <span class="impact {impact_class}">{v.get('impact', 'moderate').upper()}</span>
                </h4>
                <p><strong>Description:</strong> {v.get('description', 'No description')}</p>
                <p><strong>WCAG:</strong> {', '.join(v.get('wcag_criteria', []))}</p>
                <p><strong>Elements:</strong> {v.get('node_count', 0)} affected</p>
                {metadata_html}
                {self._format_fix(v.get('suggested_fix'))}
            </div>"""
        html += "</section>"
        return html
    
    def _format_warnings_section(self, warnings: List[Dict]) -> str:
        """Format warnings section"""
        if not warnings:
            return ""

        html = "<section class='warnings'><h2>Warnings</h2>"
        for w in warnings:
            metadata = w.get('metadata', {})

            # Build metadata display
            metadata_html = ""
            if metadata.get('breakpoint'):
                metadata_html += f"<p><strong>Breakpoint:</strong> {metadata.get('breakpoint')}px width</p>"
            if metadata.get('pseudoclass'):
                metadata_html += f"<p><strong>CSS State:</strong> {metadata.get('pseudoclass')}</p>"

            html += f"""
            <div class="warning">
                <h4>{w.get('rule_id', 'Unknown')}</h4>
                <p>{w.get('description', 'No description')}</p>
                {metadata_html}
            </div>"""
        html += "</section>"
        return html
    
    def _format_info_section(self, info_items: List[Dict]) -> str:
        """Format info section"""
        if not info_items:
            return ""

        html = "<section class='info'><h2>Information Notes</h2>"
        for item in info_items:
            metadata = item.get('metadata', {})

            # Build metadata display
            metadata_html = ""
            if metadata.get('breakpoint'):
                metadata_html += f"<p><strong>Breakpoint:</strong> {metadata.get('breakpoint')}px width</p>"
            if metadata.get('pseudoclass'):
                metadata_html += f"<p><strong>CSS State:</strong> {metadata.get('pseudoclass')}</p>"

            html += f"""
            <div class="info-item">
                <h4>{item.get('id', 'Unknown')}</h4>
                <p><strong>Description:</strong> {item.get('description', 'No description')}</p>
                <p><strong>Category:</strong> {item.get('category', 'General')}</p>
                {f"<p><strong>WCAG:</strong> {', '.join(item.get('wcag_criteria', []))}</p>" if item.get('wcag_criteria') else ""}
                {metadata_html}
            </div>"""
        html += "</section>"
        return html
    
    def _format_discovery_section(self, discovery_items: List[Dict]) -> str:
        """Format discovery section"""
        if not discovery_items:
            return ""

        html = "<section class='discovery'><h2>Discovery Items</h2>"
        html += "<p><em>These items require manual inspection to ensure accessibility.</em></p>"
        for item in discovery_items:
            metadata = item.get('metadata', {})

            # Build metadata display
            metadata_html = ""
            if metadata.get('breakpoint'):
                metadata_html += f"<p><strong>Breakpoint:</strong> {metadata.get('breakpoint')}px width</p>"
            if metadata.get('pseudoclass'):
                metadata_html += f"<p><strong>CSS State:</strong> {metadata.get('pseudoclass')}</p>"

            html += f"""
            <div class="discovery-item">
                <h4>{item.get('id', 'Unknown')}</h4>
                <p><strong>Description:</strong> {item.get('description', 'No description')}</p>
                <p><strong>Category:</strong> {item.get('category', 'General')}</p>
                {f"<p><strong>Location:</strong> <code>{item.get('xpath', 'Not specified')}</code></p>" if item.get('xpath') else ""}
                {metadata_html}
            </div>"""
        html += "</section>"
        return html
    
    def _format_ai_findings_section(self, findings: List) -> str:
        """Format AI findings section"""
        if not findings:
            return ""
        
        html = "<section class='ai-findings'><h2>AI Analysis Findings</h2>"
        for f in findings:
            severity_class = f.severity.value.lower() if hasattr(f, 'severity') else 'moderate'
            html += f"""
            <div class="ai-finding">
                <h4>{f.type if hasattr(f, 'type') else 'AI Finding'}
                    <span class="impact {severity_class}">{f.severity.value.upper() if hasattr(f, 'severity') else 'MODERATE'}</span>
                </h4>
                <p><strong>Description:</strong> {f.description if hasattr(f, 'description') else 'No description'}</p>
                <p><strong>Confidence:</strong> {f.confidence * 100 if hasattr(f, 'confidence') else 85:.0f}%</p>
                {self._format_fix(f.suggested_fix if hasattr(f, 'suggested_fix') else None)}
            </div>"""
        html += "</section>"
        return html
    
    def _format_passes_section(self, passes: List[Dict]) -> str:
        """Format passes section"""
        if not passes:
            return ""
        
        html = "<section class='passes'><h2>Passes</h2><ul>"
        for p in passes[:10]:  # Show first 10 passes
            html += f"<li>{p.get('rule_id', 'Unknown')}: {p.get('description', 'Passed')}</li>"
        if len(passes) > 10:
            html += f"<li>... and {len(passes) - 10} more</li>"
        html += "</ul></section>"
        return html
    
    def _format_fix(self, fix: str) -> str:
        """Format suggested fix"""
        if not fix:
            return ""
        return f'<p><strong>Suggested Fix:</strong> {fix}</p>'
    
    def _format_violation_types_section(self, violation_types: Dict) -> str:
        """Format violation types section"""
        html = "<section class='violation-types'><h2>Violation Types</h2><table>"
        html += "<thead><tr><th>Rule</th><th>Count</th><th>Description</th><th>Pages Affected</th></tr></thead><tbody>"
        
        sorted_types = sorted(violation_types.items(), key=lambda x: x[1]['count'], reverse=True)
        for rule_id, info in sorted_types[:20]:  # Show top 20
            pages_summary = f"{len(set(info['pages']))} pages"
            html += f"""
            <tr>
                <td>{rule_id}</td>
                <td>{info['count']}</td>
                <td>{info['description']}</td>
                <td>{pages_summary}</td>
            </tr>"""
        
        html += "</tbody></table></section>"
        return html
    
    def _format_pages_table(self, page_results: List[Dict]) -> str:
        """Format pages table"""
        html = "<section class='pages'><h2>Pages</h2><table>"
        html += "<thead><tr><th>Page</th><th>Violations</th><th>Warnings</th><th>Last Tested</th></tr></thead><tbody>"
        
        for pr in page_results:
            page = pr['page']
            test = pr['test_result']
            html += f"""
            <tr>
                <td><a href="{page.url}" target="_blank">{page.url}</a></td>
                <td class="violations">{test.violation_count}</td>
                <td class="warnings">{test.warning_count}</td>
                <td>{test.test_date}</td>
            </tr>"""
        
        html += "</tbody></table></section>"
        return html
    
    def _format_websites_section(self, websites: List[Dict]) -> str:
        """Format websites section with detailed test results"""
        html = "<section class='websites'><h2>Websites</h2>"
        
        for wd in websites:
            website = wd['website']
            pages = wd['pages']
            
            # Calculate totals for all categories
            total_violations = 0
            total_warnings = 0
            total_info = 0
            total_discovery = 0
            total_passes = 0
            
            all_violations = []
            all_warnings = []
            all_info = []
            all_discovery = []
            
            for p in pages:
                test_result = p['test_result']
                if hasattr(test_result, 'violation_count'):
                    total_violations += test_result.violation_count
                    total_warnings += test_result.warning_count
                    total_info += test_result.info_count
                    total_discovery += test_result.discovery_count
                    total_passes += test_result.pass_count
                    
                    # Collect all issues for detailed display
                    for v in test_result.violations:
                        all_violations.append({'page': p['page'].url, 'issue': v})
                    for w in test_result.warnings:
                        all_warnings.append({'page': p['page'].url, 'issue': w})
                    for i in test_result.info:
                        all_info.append({'page': p['page'].url, 'issue': i})
                    for d in test_result.discovery:
                        all_discovery.append({'page': p['page'].url, 'issue': d})
            
            # Website header with all statistics
            html += f"""
            <div class="website">
                <h3>{website.name}</h3>
                <p><a href="{website.url}" target="_blank">{website.url}</a></p>
                <div class="stats-grid" style="margin: 20px 0;">
                    <div class="stat-card violations">
                        <h4>{total_violations}</h4>
                        <p>Errors</p>
                    </div>
                    <div class="stat-card warnings">
                        <h4>{total_warnings}</h4>
                        <p>Warnings</p>
                    </div>
                    <div class="stat-card info">
                        <h4>{total_info}</h4>
                        <p>Info</p>
                    </div>
                    <div class="stat-card discovery">
                        <h4>{total_discovery}</h4>
                        <p>Discovery</p>
                    </div>
                    <div class="stat-card passes">
                        <h4>{total_passes}</h4>
                        <p>Passed</p>
                    </div>
                </div>
                
                <p><strong>Pages Tested:</strong> {len(pages)}</p>
                """
            
            # Add detailed issues if they exist
            if all_violations:
                html += "<h4>Errors (Violations)</h4><ul>"
                for item in all_violations[:10]:  # Show first 10
                    issue = item['issue']
                    html += f"<li><strong>{issue.id if hasattr(issue, 'id') else 'Unknown'}:</strong> {issue.description if hasattr(issue, 'description') else 'No description'} - <em>{item['page']}</em></li>"
                if len(all_violations) > 10:
                    html += f"<li><em>... and {len(all_violations) - 10} more errors</em></li>"
                html += "</ul>"
            
            if all_warnings:
                html += "<h4>Warnings</h4><ul>"
                for item in all_warnings[:10]:  # Show first 10
                    issue = item['issue']
                    html += f"<li><strong>{issue.id if hasattr(issue, 'id') else 'Unknown'}:</strong> {issue.description if hasattr(issue, 'description') else 'No description'} - <em>{item['page']}</em></li>"
                if len(all_warnings) > 10:
                    html += f"<li><em>... and {len(all_warnings) - 10} more warnings</em></li>"
                html += "</ul>"
            
            if all_info:
                html += "<h4>Information Notes</h4><ul>"
                for item in all_info[:5]:  # Show first 5
                    issue = item['issue']
                    html += f"<li><strong>{issue.id if hasattr(issue, 'id') else 'Unknown'}:</strong> {issue.description if hasattr(issue, 'description') else 'No description'} - <em>{item['page']}</em></li>"
                if len(all_info) > 5:
                    html += f"<li><em>... and {len(all_info) - 5} more info notes</em></li>"
                html += "</ul>"
            
            if all_discovery:
                html += "<h4>Discovery Items</h4><ul>"
                for item in all_discovery[:5]:  # Show first 5
                    issue = item['issue']
                    html += f"<li><strong>{issue.id if hasattr(issue, 'id') else 'Unknown'}:</strong> {issue.description if hasattr(issue, 'description') else 'No description'} - <em>{item['page']}</em></li>"
                if len(all_discovery) > 5:
                    html += f"<li><em>... and {len(all_discovery) - 5} more discovery items</em></li>"
                html += "</ul>"
            
            html += "</div>"
        
        html += "</section>"
        return html


class JSONFormatter(BaseFormatter):
    """JSON report formatter"""
    
    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.extension = 'json'
    
    def format_page_report(self, data: Dict[str, Any]) -> str:
        """Generate JSON report for a page"""
        return json.dumps(data, indent=2, default=str)
    
    def format_website_report(self, data: Dict[str, Any]) -> str:
        """Generate JSON report for a website"""
        return json.dumps(data, indent=2, default=str)
    
    def format_project_report(self, data: Dict[str, Any]) -> str:
        """Generate JSON report for a project"""
        return json.dumps(data, indent=2, default=str)
    
    def format_summary_report(self, data: Dict[str, Any]) -> str:
        """Generate JSON summary report"""
        return json.dumps(data, indent=2, default=str)


class CSVFormatter(BaseFormatter):
    """CSV report formatter"""
    
    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.extension = 'csv'
    
    def format_page_report(self, data: Dict[str, Any]) -> str:
        """Generate CSV report for a page"""
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Type', 'Rule ID', 'Description', 'Impact', 'WCAG', 'Elements', 'Fix'])
        
        # Write violations
        for v in data['violations']:
            writer.writerow([
                'Violation',
                v.get('rule_id', ''),
                v.get('description', ''),
                v.get('impact', ''),
                ', '.join(v.get('wcag_criteria', [])),
                v.get('node_count', 0),
                v.get('suggested_fix', '')
            ])
        
        # Write warnings
        for w in data['warnings']:
            writer.writerow([
                'Warning',
                w.get('rule_id', ''),
                w.get('description', ''),
                '',
                '',
                w.get('node_count', 0),
                ''
            ])
        
        return output.getvalue()
    
    def format_website_report(self, data: Dict[str, Any]) -> str:
        """Generate CSV report for a website"""
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Page URL', 'Violations', 'Warnings', 'Passes', 'Last Tested'])
        
        # Write page data
        for pr in data['pages']:
            page = pr['page']
            test = pr['test_result']
            writer.writerow([
                page.url,
                test.violation_count,
                test.warning_count,
                test.pass_count,
                test.test_date
            ])
        
        return output.getvalue()
    
    def format_project_report(self, data: Dict[str, Any]) -> str:
        """Generate CSV report for a project"""
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Website', 'URL', 'Pages', 'Total Violations', 'Total Warnings'])
        
        # Write website data
        for wd in data['websites']:
            website = wd['website']
            pages = wd['pages']
            
            total_violations = sum(p['test_result'].violation_count for p in pages)
            total_warnings = sum(p['test_result'].warning_count for p in pages)
            
            writer.writerow([
                website.name,
                website.url,
                len(pages),
                total_violations,
                total_warnings
            ])
        
        return output.getvalue()
    
    def format_summary_report(self, data: Dict[str, Any]) -> str:
        """Generate CSV summary report"""
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Project', 'Websites', 'Pages Tested', 'Violations'])
        
        # Write project data
        for project in data['projects']:
            writer.writerow([
                project['name'],
                project['websites'],
                project['pages_tested'],
                project['violations']
            ])
        
        return output.getvalue()


class ExcelFormatter(BaseFormatter):
    """Excel report formatter"""
    
    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.extension = 'xlsx'
        # Import openpyxl here to avoid import errors if not needed
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            self.Workbook = Workbook
            self.Font = Font
            self.Fill = Fill
            self.PatternFill = PatternFill
            self.Alignment = Alignment
            self.Border = Border
            self.Side = Side
            self.get_column_letter = get_column_letter
            self.has_openpyxl = True
        except ImportError:
            logger.warning("openpyxl not installed - Excel export will return JSON")
            self.has_openpyxl = False
    
    def _get_styles(self):
        """Get common Excel styles"""
        if not self.has_openpyxl:
            return {}
        
        return {
            'header': {
                'font': self.Font(bold=True, color="FFFFFF", size=12),
                'fill': self.PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid"),
                'alignment': self.Alignment(horizontal="center", vertical="center"),
                'border': self.Border(
                    left=self.Side(style='thin'),
                    right=self.Side(style='thin'),
                    top=self.Side(style='thin'),
                    bottom=self.Side(style='thin')
                )
            },
            'subheader': {
                'font': self.Font(bold=True, size=11),
                'fill': self.PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid"),
                'alignment': self.Alignment(horizontal="left", vertical="center")
            },
            'violation': {
                'fill': self.PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
            },
            'warning': {
                'fill': self.PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            },
            'info': {
                'fill': self.PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
            },
            'discovery': {
                'fill': self.PatternFill(start_color="E6E0FF", end_color="E6E0FF", fill_type="solid")
            },
            'pass': {
                'fill': self.PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            },
            'critical': {
                'font': self.Font(color="DC3545", bold=True)
            },
            'serious': {
                'font': self.Font(color="FD7E14", bold=True)
            },
            'moderate': {
                'font': self.Font(color="FFC107")
            },
            'minor': {
                'font': self.Font(color="6C757D")
            }
        }
    
    def format_page_report(self, data: Dict[str, Any]) -> bytes:
        """Generate Excel report for a page"""
        if not self.has_openpyxl:
            return json.dumps(data, indent=2, default=str).encode('utf-8')
        
        wb = self.Workbook()
        styles = self._get_styles()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, data, styles)

        # Multi-State Summary Sheet (if applicable)
        test_result = data.get('test_result', {})
        if test_result and test_result.get('session_id'):
            ws_states = wb.create_sheet("Page States")
            self._create_page_states_sheet(ws_states, data, styles)

        # All Issues Sheet (combined view)
        ws_all_issues = wb.create_sheet("All Issues")
        self._create_all_issues_sheet(ws_all_issues, data, styles)

        # Violations Sheet
        if data.get('violations'):
            ws_violations = wb.create_sheet("Violations")
            self._create_violations_sheet(ws_violations, data['violations'], styles)
        
        # Warnings Sheet
        if data.get('warnings'):
            ws_warnings = wb.create_sheet("Warnings")
            self._create_warnings_sheet(ws_warnings, data['warnings'], styles)
        
        # Info Sheet
        if data.get('info'):
            ws_info = wb.create_sheet("Info")
            self._create_info_sheet(ws_info, data['info'], styles)
        
        # Discovery Sheet
        if data.get('discovery'):
            ws_discovery = wb.create_sheet("Discovery")
            self._create_discovery_sheet(ws_discovery, data['discovery'], styles)
        
        # AI Findings Sheet
        if data.get('ai_findings'):
            ws_ai = wb.create_sheet("AI Findings")
            self._create_ai_findings_sheet(ws_ai, data['ai_findings'], styles)
        
        # Passes Sheet
        if data.get('passes'):
            ws_passes = wb.create_sheet("Passes")
            self._create_passes_sheet(ws_passes, data['passes'], styles)
        
        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def format_website_report(self, data: Dict[str, Any]) -> bytes:
        """Generate Excel report for a website"""
        if not self.has_openpyxl:
            return json.dumps(data, indent=2, default=str).encode('utf-8')
        
        wb = self.Workbook()
        styles = self._get_styles()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Website Summary"
        self._create_website_summary_sheet(ws_summary, data, styles)
        
        # Pages Sheet
        if data.get('pages'):
            ws_pages = wb.create_sheet("Pages")
            self._create_pages_sheet(ws_pages, data['pages'], styles)
        
        # Violation Types Sheet
        if data.get('violation_types'):
            ws_types = wb.create_sheet("Violation Types")
            self._create_violation_types_sheet(ws_types, data['violation_types'], styles)
        
        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def format_project_report(self, data: Dict[str, Any]) -> bytes:
        """Generate Excel report for a project"""
        if not self.has_openpyxl:
            return json.dumps(data, indent=2, default=str).encode('utf-8')
        
        wb = self.Workbook()
        styles = self._get_styles()
        
        # Project Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Project Summary"
        self._create_project_summary_sheet(ws_summary, data, styles)
        
        # Websites Sheet
        if data.get('websites'):
            ws_websites = wb.create_sheet("Websites")
            self._create_websites_sheet(ws_websites, data['websites'], styles)

        # All Issues Sheet (combined view across all pages)
        ws_all_issues = wb.create_sheet("All Issues")
        self._create_project_all_issues_sheet(ws_all_issues, data, styles)

        # All Issues (Deduplicated) Sheet - groups by common components
        ws_deduped_issues = wb.create_sheet("All Issues (Deduplicated)")
        self._create_project_deduped_issues_sheet(ws_deduped_issues, data, styles)

        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def format_all_projects_report(self, data: Dict[str, Any]) -> bytes:
        """Generate Excel report for all projects"""
        if not self.has_openpyxl:
            return json.dumps(data, indent=2, default=str).encode('utf-8')
        
        wb = self.Workbook()
        styles = self._get_styles()
        
        # Overall Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Overall Summary"
        self._create_all_projects_summary_sheet(ws_summary, data, styles)
        
        # Projects Breakdown Sheet
        if data.get('projects'):
            ws_projects = wb.create_sheet("Projects Breakdown")
            self._create_projects_breakdown_sheet(ws_projects, data['projects'], styles)
        
        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def format_summary_report(self, data: Dict[str, Any]) -> bytes:
        """Generate Excel summary report"""
        if not self.has_openpyxl:
            return json.dumps(data, indent=2, default=str).encode('utf-8')
        
        wb = self.Workbook()
        styles = self._get_styles()
        
        # Executive Summary Sheet
        ws = wb.active
        ws.title = "Executive Summary"
        
        # Headers
        headers = ['Project', 'Websites', 'Pages Tested', 'Violations', 'Warnings']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, styles['header'])
        
        # Data
        row = 2
        for project in data.get('projects', []):
            ws.cell(row=row, column=1, value=project.get('name', ''))
            ws.cell(row=row, column=2, value=project.get('websites', 0))
            ws.cell(row=row, column=3, value=project.get('pages_tested', 0))
            
            violations_cell = ws.cell(row=row, column=4, value=project.get('violations', 0))
            if project.get('violations', 0) > 0:
                violations_cell.fill = styles['violation']['fill']
            
            warnings_cell = ws.cell(row=row, column=5, value=project.get('warnings', 0))
            if project.get('warnings', 0) > 0:
                warnings_cell.fill = styles['warning']['fill']
            
            row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _create_summary_sheet(self, ws, data, styles):
        """Create page summary sheet"""
        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"Accessibility Report - {data.get('page', {}).get('url', 'Unknown')}"
        title_cell.font = self.Font(bold=True, size=14)
        title_cell.alignment = self.Alignment(horizontal="center")
        
        # Metadata
        row = 3
        ws.cell(row=row, column=1, value="Website:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=data.get('website', {}).get('name', ''))
        row += 1
        ws.cell(row=row, column=1, value="Project:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=data.get('project', {}).get('name', ''))
        row += 1
        ws.cell(row=row, column=1, value="Generated:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=data.get('generated_at', ''))
        
        # Statistics
        row = 7
        ws.cell(row=row, column=1, value="Summary Statistics").font = self.Font(bold=True, size=12)
        row += 1
        
        stats = data.get('statistics', {})
        ws.cell(row=row, column=1, value="Violations:")
        ws.cell(row=row, column=2, value=stats.get('violations', 0))
        if stats.get('violations', 0) > 0:
            ws.cell(row=row, column=2).fill = styles['violation']['fill']
        row += 1
        
        ws.cell(row=row, column=1, value="Warnings:")
        ws.cell(row=row, column=2, value=stats.get('warnings', 0))
        if stats.get('warnings', 0) > 0:
            ws.cell(row=row, column=2).fill = styles['warning']['fill']
        row += 1
        
        ws.cell(row=row, column=1, value="Passes:")
        ws.cell(row=row, column=2, value=stats.get('passes', 0))
        ws.cell(row=row, column=2).fill = styles['pass']['fill']
        row += 1
        
        ws.cell(row=row, column=1, value="Test Duration (ms):")
        ws.cell(row=row, column=2, value=stats.get('duration_ms', 0))
        
        self._auto_adjust_columns(ws)
    
    def _create_violations_sheet(self, ws, violations, styles):
        """Create violations sheet"""
        headers = ['Rule ID', 'Description', 'Impact', 'WCAG Criteria', 'Elements Affected', 'Suggested Fix', 'Test User', 'User Roles']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, styles['header'])

        row = 2
        for v in violations:
            ws.cell(row=row, column=1, value=v.get('rule_id', ''))
            ws.cell(row=row, column=2, value=v.get('description', ''))

            impact = v.get('impact', 'moderate')
            impact_cell = ws.cell(row=row, column=3, value=impact.upper())
            if impact.lower() in styles:
                impact_cell.font = styles[impact.lower()]['font']

            ws.cell(row=row, column=4, value=', '.join(v.get('wcag_criteria', [])))
            ws.cell(row=row, column=5, value=v.get('node_count', 0))
            ws.cell(row=row, column=6, value=v.get('suggested_fix', ''))

            # Add authenticated user info
            metadata = v.get('metadata', {})
            auth_user = metadata.get('authenticated_user', {})
            if auth_user:
                ws.cell(row=row, column=7, value=auth_user.get('display_name', ''))
                ws.cell(row=row, column=8, value=', '.join(auth_user.get('roles', [])))
            else:
                ws.cell(row=row, column=7, value='Guest')
                ws.cell(row=row, column=8, value='no login')

            # Apply row coloring
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = styles['violation']['fill']

            row += 1

        self._auto_adjust_columns(ws)
    
    def _create_warnings_sheet(self, ws, warnings, styles):
        """Create warnings sheet"""
        headers = ['Rule ID', 'Description', 'Elements Affected', 'Test User', 'User Roles']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, styles['header'])

        row = 2
        for w in warnings:
            ws.cell(row=row, column=1, value=w.get('rule_id', ''))
            ws.cell(row=row, column=2, value=w.get('description', ''))
            ws.cell(row=row, column=3, value=w.get('node_count', 0))

            # Add authenticated user info
            metadata = w.get('metadata', {})
            auth_user = metadata.get('authenticated_user', {})
            if auth_user:
                ws.cell(row=row, column=4, value=auth_user.get('display_name', ''))
                ws.cell(row=row, column=5, value=', '.join(auth_user.get('roles', [])))
            else:
                ws.cell(row=row, column=4, value='Guest')
                ws.cell(row=row, column=5, value='no login')

            # Apply row coloring
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = styles['warning']['fill']

            row += 1

        self._auto_adjust_columns(ws)
    
    def _create_info_sheet(self, ws, info_items, styles):
        """Create info sheet"""
        headers = ['ID', 'Description', 'Category', 'WCAG Criteria', 'Location', 'Test User', 'User Roles']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, styles['header'])

        row = 2
        for item in info_items:
            ws.cell(row=row, column=1, value=item.get('id', ''))
            ws.cell(row=row, column=2, value=item.get('description', ''))
            ws.cell(row=row, column=3, value=item.get('category', ''))
            ws.cell(row=row, column=4, value=', '.join(item.get('wcag_criteria', [])))
            ws.cell(row=row, column=5, value=item.get('xpath', ''))

            # Add authenticated user info
            metadata = item.get('metadata', {})
            auth_user = metadata.get('authenticated_user', {})
            if auth_user:
                ws.cell(row=row, column=6, value=auth_user.get('display_name', ''))
                ws.cell(row=row, column=7, value=', '.join(auth_user.get('roles', [])))
            else:
                ws.cell(row=row, column=6, value='Guest')
                ws.cell(row=row, column=7, value='no login')

            # Apply info coloring (blue)
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = styles['info']['fill']

            row += 1

        self._auto_adjust_columns(ws)
    
    def _create_discovery_sheet(self, ws, discovery_items, styles):
        """Create discovery sheet"""
        headers = ['ID', 'Description', 'Category', 'Location', 'Manual Check Required', 'Test User', 'User Roles']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, styles['header'])

        row = 2
        for item in discovery_items:
            ws.cell(row=row, column=1, value=item.get('id', ''))
            ws.cell(row=row, column=2, value=item.get('description', ''))
            ws.cell(row=row, column=3, value=item.get('category', ''))
            ws.cell(row=row, column=4, value=item.get('xpath', ''))
            ws.cell(row=row, column=5, value='Yes')

            # Add authenticated user info
            metadata = item.get('metadata', {})
            auth_user = metadata.get('authenticated_user', {})
            if auth_user:
                ws.cell(row=row, column=6, value=auth_user.get('display_name', ''))
                ws.cell(row=row, column=7, value=', '.join(auth_user.get('roles', [])))
            else:
                ws.cell(row=row, column=6, value='Guest')
                ws.cell(row=row, column=7, value='no login')

            # Apply discovery coloring (purple)
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                if 'discovery' in styles:
                    cell.fill = styles['discovery']['fill']
                else:
                    # Use a purple-ish color if not defined
                    from openpyxl.styles import PatternFill
                    cell.fill = PatternFill(start_color="E6E0FF", end_color="E6E0FF", fill_type="solid")

            row += 1

        self._auto_adjust_columns(ws)
    
    def _create_ai_findings_sheet(self, ws, findings, styles):
        """Create AI findings sheet"""
        headers = ['Type', 'Description', 'Severity', 'Confidence', 'Suggested Fix']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, styles['header'])
        
        row = 2
        for f in findings:
            ws.cell(row=row, column=1, value=getattr(f, 'type', 'AI Finding'))
            ws.cell(row=row, column=2, value=getattr(f, 'description', ''))
            
            severity = getattr(f, 'severity', None)
            if severity:
                severity_value = severity.value if hasattr(severity, 'value') else str(severity)
                severity_cell = ws.cell(row=row, column=3, value=severity_value.upper())
                if severity_value.lower() in styles:
                    severity_cell.font = styles[severity_value.lower()]['font']
            
            confidence = getattr(f, 'confidence', 0.85)
            ws.cell(row=row, column=4, value=f"{confidence * 100:.0f}%")
            ws.cell(row=row, column=5, value=getattr(f, 'suggested_fix', ''))
            
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_passes_sheet(self, ws, passes, styles):
        """Create passes sheet"""
        headers = ['Rule ID', 'Description']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, styles['header'])
        
        row = 2
        for p in passes:
            ws.cell(row=row, column=1, value=p.get('rule_id', ''))
            ws.cell(row=row, column=2, value=p.get('description', 'Passed'))
            
            # Apply row coloring
            for col in range(1, 3):
                ws.cell(row=row, column=col).fill = styles['pass']['fill']
            
            row += 1
        
        self._auto_adjust_columns(ws)

    def _create_all_issues_sheet(self, ws, data, styles):
        """Create a combined sheet with all issues (violations, warnings, info, discovery)"""
        headers = ['Type', 'Impact', 'Rule ID', 'Touchpoint', 'What', 'Why Important', 'Who Affected', 'How to Remediate', 'WCAG Criteria', 'Location (XPath)', 'Element', 'Page URL', 'Breakpoint (px)', 'Pseudoclass', 'Page State', 'Test User', 'User Roles']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, styles['header'])

        row = 2

        # Get page state description from test result if available
        page_state_desc = ''
        test_result = data.get('test_result', {})
        if test_result:
            page_state = test_result.get('page_state')
            if page_state:
                if isinstance(page_state, dict):
                    page_state_desc = page_state.get('description', '')
                elif hasattr(page_state, 'description'):
                    page_state_desc = page_state.description

        # Add all violations
        for v in data.get('violations', []):
            ws.cell(row=row, column=1, value='Violation')
            ws.cell(row=row, column=2, value=v.get('impact', 'Unknown').upper())
            ws.cell(row=row, column=3, value=v.get('id', v.get('rule_id', '')))
            ws.cell(row=row, column=4, value=v.get('touchpoint', v.get('category', '')))
            ws.cell(row=row, column=5, value=v.get('description_full', v.get('what', v.get('description', ''))))
            ws.cell(row=row, column=6, value=v.get('why_it_matters', ''))
            ws.cell(row=row, column=7, value=v.get('who_it_affects', ''))
            ws.cell(row=row, column=8, value=v.get('how_to_fix', v.get('remediation', v.get('suggested_fix', ''))))
            ws.cell(row=row, column=9, value=v.get('wcag_full', ', '.join(v.get('wcag_criteria', [])) if isinstance(v.get('wcag_criteria'), list) else v.get('wcag_criteria', '')))
            ws.cell(row=row, column=10, value=v.get('xpath', ''))
            ws.cell(row=row, column=11, value=v.get('element', ''))
            ws.cell(row=row, column=12, value=v.get('url', ''))

            # Add metadata columns
            metadata = v.get('metadata', {})
            ws.cell(row=row, column=13, value=metadata.get('breakpoint', ''))
            ws.cell(row=row, column=14, value=metadata.get('pseudoclass', ''))
            ws.cell(row=row, column=15, value=page_state_desc)

            # Add authenticated user info
            auth_user = metadata.get('authenticated_user', {})
            if auth_user:
                ws.cell(row=row, column=16, value=auth_user.get('display_name', ''))
                ws.cell(row=row, column=17, value=', '.join(auth_user.get('roles', [])))
            else:
                ws.cell(row=row, column=16, value='Guest')
                ws.cell(row=row, column=17, value='no login')

            # Apply violation coloring
            for col in range(1, 18):
                ws.cell(row=row, column=col).fill = styles['violation']['fill']

            row += 1

        # Add all warnings
        for w in data.get('warnings', []):
            ws.cell(row=row, column=1, value='Warning')
            ws.cell(row=row, column=2, value=w.get('impact', 'MEDIUM').upper())
            ws.cell(row=row, column=3, value=w.get('id', w.get('rule_id', '')))
            ws.cell(row=row, column=4, value=w.get('touchpoint', w.get('category', '')))
            ws.cell(row=row, column=5, value=w.get('description_full', w.get('what', w.get('description', ''))))
            ws.cell(row=row, column=6, value=w.get('why_it_matters', ''))
            ws.cell(row=row, column=7, value=w.get('who_it_affects', ''))
            ws.cell(row=row, column=8, value=w.get('how_to_fix', w.get('remediation', w.get('suggested_fix', ''))))
            ws.cell(row=row, column=9, value=w.get('wcag_full', ', '.join(w.get('wcag_criteria', [])) if isinstance(w.get('wcag_criteria'), list) else w.get('wcag_criteria', '')))
            ws.cell(row=row, column=10, value=w.get('xpath', ''))
            ws.cell(row=row, column=11, value=w.get('element', ''))
            ws.cell(row=row, column=12, value=w.get('url', ''))

            # Add metadata columns
            metadata = w.get('metadata', {})
            ws.cell(row=row, column=13, value=metadata.get('breakpoint', ''))
            ws.cell(row=row, column=14, value=metadata.get('pseudoclass', ''))
            ws.cell(row=row, column=15, value=page_state_desc)

            # Add authenticated user info
            auth_user = metadata.get('authenticated_user', {})
            if auth_user:
                ws.cell(row=row, column=16, value=auth_user.get('display_name', ''))
                ws.cell(row=row, column=17, value=', '.join(auth_user.get('roles', [])))
            else:
                ws.cell(row=row, column=16, value='Guest')
                ws.cell(row=row, column=17, value='no login')

            # Apply warning coloring
            for col in range(1, 18):
                ws.cell(row=row, column=col).fill = styles['warning']['fill']

            row += 1

        # Add all info items
        for i in data.get('info', []):
            ws.cell(row=row, column=1, value='Info')
            ws.cell(row=row, column=2, value='INFO')
            ws.cell(row=row, column=3, value=i.get('id', ''))
            ws.cell(row=row, column=4, value=i.get('touchpoint', i.get('category', '')))
            ws.cell(row=row, column=5, value=i.get('description_full', i.get('what', i.get('description', ''))))
            ws.cell(row=row, column=6, value=i.get('why_it_matters', ''))
            ws.cell(row=row, column=7, value=i.get('who_it_affects', ''))
            ws.cell(row=row, column=8, value=i.get('how_to_fix', i.get('remediation', '')))
            ws.cell(row=row, column=9, value=i.get('wcag_full', ', '.join(i.get('wcag_criteria', [])) if isinstance(i.get('wcag_criteria'), list) else i.get('wcag_criteria', '')))
            ws.cell(row=row, column=10, value=i.get('xpath', ''))
            ws.cell(row=row, column=11, value=i.get('element', ''))
            ws.cell(row=row, column=12, value=i.get('url', ''))

            # Add metadata columns
            metadata = i.get('metadata', {})
            ws.cell(row=row, column=13, value=metadata.get('breakpoint', ''))
            ws.cell(row=row, column=14, value=metadata.get('pseudoclass', ''))
            ws.cell(row=row, column=15, value=page_state_desc)

            # Add authenticated user info
            auth_user = metadata.get('authenticated_user', {})
            if auth_user:
                ws.cell(row=row, column=16, value=auth_user.get('display_name', ''))
                ws.cell(row=row, column=17, value=', '.join(auth_user.get('roles', [])))
            else:
                ws.cell(row=row, column=16, value='Guest')
                ws.cell(row=row, column=17, value='no login')

            # Apply info coloring
            for col in range(1, 18):
                ws.cell(row=row, column=col).fill = styles['info']['fill']

            row += 1

        # Add all discovery items
        for d in data.get('discovery', []):
            ws.cell(row=row, column=1, value='Discovery')
            ws.cell(row=row, column=2, value='DISCOVERY')
            ws.cell(row=row, column=3, value=d.get('id', d.get('err', '')))
            ws.cell(row=row, column=4, value=d.get('touchpoint', d.get('category', '')))
            ws.cell(row=row, column=5, value=d.get('description_full', d.get('what', d.get('description', ''))))
            ws.cell(row=row, column=6, value=d.get('why_it_matters', ''))
            ws.cell(row=row, column=7, value=d.get('who_it_affects', ''))
            ws.cell(row=row, column=8, value=d.get('how_to_fix', d.get('remediation', '')))
            ws.cell(row=row, column=9, value=d.get('wcag_full', ', '.join(d.get('wcag_criteria', [])) if isinstance(d.get('wcag_criteria'), list) else d.get('wcag_criteria', '')))
            ws.cell(row=row, column=10, value=d.get('xpath', ''))
            ws.cell(row=row, column=11, value=d.get('element', ''))
            ws.cell(row=row, column=12, value=d.get('url', ''))

            # Add metadata columns
            metadata = d.get('metadata', {})
            ws.cell(row=row, column=13, value=metadata.get('breakpoint', ''))
            ws.cell(row=row, column=14, value=metadata.get('pseudoclass', ''))
            ws.cell(row=row, column=15, value=page_state_desc)

            # Add authenticated user info
            auth_user = metadata.get('authenticated_user', {})
            if auth_user:
                ws.cell(row=row, column=16, value=auth_user.get('display_name', ''))
                ws.cell(row=row, column=17, value=', '.join(auth_user.get('roles', [])))
            else:
                ws.cell(row=row, column=16, value='Guest')
                ws.cell(row=row, column=17, value='no login')

            # Apply discovery coloring
            for col in range(1, 18):
                cell = ws.cell(row=row, column=col)
                if 'discovery' in styles:
                    cell.fill = styles['discovery']['fill']
                else:
                    from openpyxl.styles import PatternFill
                    cell.fill = PatternFill(start_color="E6E0FF", end_color="E6E0FF", fill_type="solid")

            row += 1

        # Add AI findings if present
        for f in data.get('ai_findings', []):
            ws.cell(row=row, column=1, value='AI Finding')
            ws.cell(row=row, column=2, value=getattr(f, 'severity', 'MEDIUM').upper())
            ws.cell(row=row, column=3, value=getattr(f, 'type', ''))
            ws.cell(row=row, column=4, value='')  # AI findings don't have touchpoint
            ws.cell(row=row, column=5, value=getattr(f, 'description', ''))
            ws.cell(row=row, column=6, value='')
            ws.cell(row=row, column=7, value='')
            ws.cell(row=row, column=8, value=getattr(f, 'suggested_fix', ''))
            ws.cell(row=row, column=9, value='')
            ws.cell(row=row, column=10, value='')
            ws.cell(row=row, column=11, value='')
            ws.cell(row=row, column=12, value=getattr(f, 'url', ''))

            # Apply AI finding coloring (use info style)
            for col in range(1, 13):
                ws.cell(row=row, column=col).fill = styles['info']['fill']

            row += 1

        self._auto_adjust_columns(ws)

    def _create_project_all_issues_sheet(self, ws, data, styles):
        """Create a combined sheet with all issues from all pages across all websites"""
        headers = ['Type', 'Impact', 'Rule ID', 'Touchpoint', 'What', 'Why Important', 'Who Affected', 'How to Remediate', 'WCAG Criteria', 'Location (XPath)', 'Element', 'Page URL', 'Website', 'Breakpoint (px)', 'Pseudoclass', 'Page State', 'Test User', 'User Roles']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, styles['header'])

        row = 2

        # Iterate through all websites and their pages
        for website_data in data.get('websites', []):
            website = website_data.get('website', {})
            website_name = website.get('name', '') if isinstance(website, dict) else getattr(website, 'name', '')

            for page_result in website_data.get('pages', []):
                page = page_result.get('page', {})
                page_url = page.get('url', '') if isinstance(page, dict) else getattr(page, 'url', '')

                test_result = page_result.get('test_result')
                if not test_result:
                    continue

                # Get page state description from test result if available
                page_state_desc = ''
                if test_result:
                    page_state = getattr(test_result, 'page_state', None) if hasattr(test_result, 'page_state') else test_result.get('page_state')
                    if page_state:
                        if isinstance(page_state, dict):
                            page_state_desc = page_state.get('description', '')
                        elif hasattr(page_state, 'description'):
                            page_state_desc = page_state.description

                # Add violations (from list attributes) - enrich with catalog data
                violations = getattr(test_result, 'violations', []) if hasattr(test_result, 'violations') else []
                for v in violations:
                    # Handle both Violation objects and dicts
                    if hasattr(v, 'to_dict'):
                        v_dict = v.to_dict()
                    else:
                        v_dict = v if isinstance(v, dict) else {}

                    # Enrich with catalog information
                    v_dict = IssueCatalog.enrich_issue(v_dict)

                    ws.cell(row=row, column=1, value='Violation')
                    ws.cell(row=row, column=2, value=str(v_dict.get('impact', 'Unknown')).upper())
                    ws.cell(row=row, column=3, value=v_dict.get('id', ''))
                    ws.cell(row=row, column=4, value=v_dict.get('touchpoint', v_dict.get('category', '')))
                    ws.cell(row=row, column=5, value=v_dict.get('description_full', v_dict.get('what', v_dict.get('description', ''))))
                    ws.cell(row=row, column=6, value=v_dict.get('why_it_matters', ''))
                    ws.cell(row=row, column=7, value=v_dict.get('who_it_affects', ''))
                    ws.cell(row=row, column=8, value=v_dict.get('how_to_fix', v_dict.get('remediation', v_dict.get('suggested_fix', ''))))
                    ws.cell(row=row, column=9, value=v_dict.get('wcag_full', ', '.join(v_dict.get('wcag_criteria', [])) if isinstance(v_dict.get('wcag_criteria'), list) else v_dict.get('wcag_criteria', '')))
                    ws.cell(row=row, column=10, value=v_dict.get('xpath', ''))
                    ws.cell(row=row, column=11, value=v_dict.get('element', ''))
                    ws.cell(row=row, column=12, value=page_url)
                    ws.cell(row=row, column=13, value=website_name)

                    # Add metadata columns
                    metadata = v_dict.get('metadata', {})
                    ws.cell(row=row, column=14, value=metadata.get('breakpoint', ''))
                    ws.cell(row=row, column=15, value=metadata.get('pseudoclass', ''))
                    ws.cell(row=row, column=16, value=page_state_desc)

                    # Add authenticated user info
                    auth_user = metadata.get('authenticated_user', {})
                    if auth_user:
                        ws.cell(row=row, column=17, value=auth_user.get('display_name', ''))
                        ws.cell(row=row, column=18, value=', '.join(auth_user.get('roles', [])))
                    else:
                        ws.cell(row=row, column=17, value='Guest')
                        ws.cell(row=row, column=18, value='no login')

                    for col in range(1, 19):
                        ws.cell(row=row, column=col).fill = styles['violation']['fill']
                    row += 1

                # Add warnings - enrich with catalog data
                warnings = getattr(test_result, 'warnings', []) if hasattr(test_result, 'warnings') else []
                for w in warnings:
                    if hasattr(w, 'to_dict'):
                        w_dict = w.to_dict()
                    else:
                        w_dict = w if isinstance(w, dict) else {}

                    # Enrich with catalog information
                    w_dict = IssueCatalog.enrich_issue(w_dict)

                    ws.cell(row=row, column=1, value='Warning')
                    ws.cell(row=row, column=2, value=str(w_dict.get('impact', 'Moderate')).upper())
                    ws.cell(row=row, column=3, value=w_dict.get('id', ''))
                    ws.cell(row=row, column=4, value=w_dict.get('touchpoint', w_dict.get('category', '')))
                    ws.cell(row=row, column=5, value=w_dict.get('description_full', w_dict.get('what', w_dict.get('description', ''))))
                    ws.cell(row=row, column=6, value=w_dict.get('why_it_matters', ''))
                    ws.cell(row=row, column=7, value=w_dict.get('who_it_affects', ''))
                    ws.cell(row=row, column=8, value=w_dict.get('how_to_fix', w_dict.get('remediation', w_dict.get('suggested_fix', ''))))
                    ws.cell(row=row, column=9, value=w_dict.get('wcag_full', ', '.join(w_dict.get('wcag_criteria', [])) if isinstance(w_dict.get('wcag_criteria'), list) else w_dict.get('wcag_criteria', '')))
                    ws.cell(row=row, column=10, value=w_dict.get('xpath', ''))
                    ws.cell(row=row, column=11, value=w_dict.get('element', ''))
                    ws.cell(row=row, column=12, value=page_url)
                    ws.cell(row=row, column=13, value=website_name)

                    # Add metadata columns
                    metadata = w_dict.get('metadata', {})
                    ws.cell(row=row, column=14, value=metadata.get('breakpoint', ''))
                    ws.cell(row=row, column=15, value=metadata.get('pseudoclass', ''))
                    ws.cell(row=row, column=16, value=page_state_desc)

                    # Add authenticated user info
                    auth_user = metadata.get('authenticated_user', {})
                    if auth_user:
                        ws.cell(row=row, column=17, value=auth_user.get('display_name', ''))
                        ws.cell(row=row, column=18, value=', '.join(auth_user.get('roles', [])))
                    else:
                        ws.cell(row=row, column=17, value='Guest')
                        ws.cell(row=row, column=18, value='no login')

                    for col in range(1, 19):
                        ws.cell(row=row, column=col).fill = styles['warning']['fill']
                    row += 1

                # Add info items - enrich with catalog data
                info_items = getattr(test_result, 'info', []) if hasattr(test_result, 'info') else []
                for i in info_items:
                    if hasattr(i, 'to_dict'):
                        i_dict = i.to_dict()
                    else:
                        i_dict = i if isinstance(i, dict) else {}

                    # Enrich with catalog information
                    i_dict = IssueCatalog.enrich_issue(i_dict)

                    ws.cell(row=row, column=1, value='Info')
                    ws.cell(row=row, column=2, value='INFO')
                    ws.cell(row=row, column=3, value=i_dict.get('id', ''))
                    ws.cell(row=row, column=4, value=i_dict.get('touchpoint', i_dict.get('category', '')))
                    ws.cell(row=row, column=5, value=i_dict.get('description_full', i_dict.get('what', i_dict.get('description', ''))))
                    ws.cell(row=row, column=6, value=i_dict.get('why_it_matters', ''))
                    ws.cell(row=row, column=7, value=i_dict.get('who_it_affects', ''))
                    ws.cell(row=row, column=8, value=i_dict.get('how_to_fix', i_dict.get('remediation', '')))
                    ws.cell(row=row, column=9, value=i_dict.get('wcag_full', ', '.join(i_dict.get('wcag_criteria', [])) if isinstance(i_dict.get('wcag_criteria'), list) else i_dict.get('wcag_criteria', '')))
                    ws.cell(row=row, column=10, value=i_dict.get('xpath', ''))
                    ws.cell(row=row, column=11, value=i_dict.get('element', ''))
                    ws.cell(row=row, column=12, value=page_url)
                    ws.cell(row=row, column=13, value=website_name)

                    # Add metadata columns
                    metadata = i_dict.get('metadata', {})
                    ws.cell(row=row, column=14, value=metadata.get('breakpoint', ''))
                    ws.cell(row=row, column=15, value=metadata.get('pseudoclass', ''))
                    ws.cell(row=row, column=16, value=page_state_desc)

                    # Add authenticated user info
                    auth_user = metadata.get('authenticated_user', {})
                    if auth_user:
                        ws.cell(row=row, column=17, value=auth_user.get('display_name', ''))
                        ws.cell(row=row, column=18, value=', '.join(auth_user.get('roles', [])))
                    else:
                        ws.cell(row=row, column=17, value='Guest')
                        ws.cell(row=row, column=18, value='no login')

                    for col in range(1, 19):
                        ws.cell(row=row, column=col).fill = styles['info']['fill']
                    row += 1

                # Add discovery items - enrich with catalog data
                discovery_items = getattr(test_result, 'discovery', []) if hasattr(test_result, 'discovery') else []
                for d in discovery_items:
                    if hasattr(d, 'to_dict'):
                        d_dict = d.to_dict()
                    else:
                        d_dict = d if isinstance(d, dict) else {}

                    # Enrich with catalog information
                    d_dict = IssueCatalog.enrich_issue(d_dict)

                    ws.cell(row=row, column=1, value='Discovery')
                    ws.cell(row=row, column=2, value='DISCOVERY')
                    ws.cell(row=row, column=3, value=d_dict.get('id', ''))
                    ws.cell(row=row, column=4, value=d_dict.get('touchpoint', d_dict.get('category', '')))
                    ws.cell(row=row, column=5, value=d_dict.get('description_full', d_dict.get('what', d_dict.get('description', ''))))
                    ws.cell(row=row, column=6, value=d_dict.get('why_it_matters', ''))
                    ws.cell(row=row, column=7, value=d_dict.get('who_it_affects', ''))
                    ws.cell(row=row, column=8, value=d_dict.get('how_to_fix', d_dict.get('remediation', '')))
                    ws.cell(row=row, column=9, value=d_dict.get('wcag_full', ', '.join(d_dict.get('wcag_criteria', [])) if isinstance(d_dict.get('wcag_criteria'), list) else d_dict.get('wcag_criteria', '')))
                    ws.cell(row=row, column=10, value=d_dict.get('xpath', ''))
                    ws.cell(row=row, column=11, value=d_dict.get('element', ''))
                    ws.cell(row=row, column=12, value=page_url)
                    ws.cell(row=row, column=13, value=website_name)

                    # Add metadata columns
                    metadata = d_dict.get('metadata', {})
                    ws.cell(row=row, column=14, value=metadata.get('breakpoint', ''))
                    ws.cell(row=row, column=15, value=metadata.get('pseudoclass', ''))
                    ws.cell(row=row, column=16, value=page_state_desc)

                    # Add authenticated user info
                    auth_user = metadata.get('authenticated_user', {})
                    if auth_user:
                        ws.cell(row=row, column=17, value=auth_user.get('display_name', ''))
                        ws.cell(row=row, column=18, value=', '.join(auth_user.get('roles', [])))
                    else:
                        ws.cell(row=row, column=17, value='Guest')
                        ws.cell(row=row, column=18, value='no login')

                    for col in range(1, 19):
                        ws.cell(row=row, column=col).fill = styles['discovery']['fill']
                    row += 1

                # Add AI findings if available
                ai_findings = getattr(test_result, 'ai_findings', []) if hasattr(test_result, 'ai_findings') else []
                for f in ai_findings:
                    if hasattr(f, 'to_dict'):
                        f_dict = f.to_dict()
                    else:
                        f_dict = f if isinstance(f, dict) else {}

                    ws.cell(row=row, column=1, value='AI Finding')
                    ws.cell(row=row, column=2, value=str(f_dict.get('severity', 'Unknown')).upper())
                    ws.cell(row=row, column=3, value=f_dict.get('type', ''))
                    ws.cell(row=row, column=4, value='')  # AI findings don't have touchpoint
                    ws.cell(row=row, column=5, value=f_dict.get('description', ''))
                    ws.cell(row=row, column=6, value='')
                    ws.cell(row=row, column=7, value='')
                    ws.cell(row=row, column=8, value=f_dict.get('suggested_fix', ''))
                    ws.cell(row=row, column=9, value='')
                    ws.cell(row=row, column=10, value='')
                    ws.cell(row=row, column=11, value='')
                    ws.cell(row=row, column=12, value=page_url)
                    ws.cell(row=row, column=13, value=website_name)
                    ws.cell(row=row, column=14, value='')  # AI findings don't have breakpoint
                    ws.cell(row=row, column=15, value='')  # AI findings don't have pseudoclass
                    ws.cell(row=row, column=16, value=page_state_desc)
                    ws.cell(row=row, column=17, value='')  # AI findings don't have test user
                    ws.cell(row=row, column=18, value='')  # AI findings don't have user roles

                    for col in range(1, 19):
                        ws.cell(row=row, column=col).fill = styles['info']['fill']
                    row += 1

        self._auto_adjust_columns(ws)

    def _xpath_is_within(self, issue_xpath: str, component_xpath: str) -> bool:
        """
        Check if an issue's XPath is within a component's XPath.

        Args:
            issue_xpath: XPath of the issue
            component_xpath: XPath of the component

        Returns:
            True if issue_xpath is within or equal to component_xpath
        """
        if not issue_xpath or not component_xpath:
            return False

        # Normalize xpaths by removing trailing slashes
        issue_xpath = issue_xpath.rstrip('/')
        component_xpath = component_xpath.rstrip('/')

        # Check if issue xpath starts with component xpath
        # e.g., /html/body/nav/a is within /html/body/nav
        return issue_xpath == component_xpath or issue_xpath.startswith(component_xpath + '/')

    def _extract_common_components(self, data: Dict[str, Any]) -> Dict[str, Dict]:
        """
        Extract common components (forms, navs, asides, sections, headers) from discovery issues.

        Args:
            data: Project report data

        Returns:
            Dictionary mapping signature -> component info with xpaths per page
        """
        common_components = {}

        # Iterate through all websites and pages
        for website_data in data.get('websites', []):
            website = website_data.get('website', {})
            website_name = website.get('name', '') if isinstance(website, dict) else getattr(website, 'name', '')

            for page_result in website_data.get('pages', []):
                page = page_result.get('page', {})
                page_url = page.get('url', '') if isinstance(page, dict) else getattr(page, 'url', '')

                test_result = page_result.get('test_result')
                if not test_result:
                    continue

                # Get discovery items
                discovery_items = getattr(test_result, 'discovery', []) if hasattr(test_result, 'discovery') else []

                for d in discovery_items:
                    if hasattr(d, 'to_dict'):
                        d_dict = d.to_dict()
                    else:
                        d_dict = d if isinstance(d, dict) else {}

                    issue_id = d_dict.get('id', '')
                    metadata = d_dict.get('metadata', {})

                    # Extract signature and xpath for different component types
                    signature = None
                    component_type = None
                    label = None

                    if issue_id in ['DiscoFormOnPage', 'forms_DiscoFormOnPage']:
                        signature = metadata.get('formSignature')
                        component_type = 'Form'
                        field_count = metadata.get('fieldCount', 0)
                        label = f"Form ({field_count} fields)"
                    elif issue_id in ['DiscoNavFound', 'landmarks_DiscoNavFound']:
                        signature = metadata.get('navSignature')
                        component_type = 'Navigation'
                        label = metadata.get('navLabel', 'Navigation')
                    elif issue_id in ['DiscoAsideFound', 'landmarks_DiscoAsideFound']:
                        signature = metadata.get('asideSignature')
                        component_type = 'Aside'
                        label = metadata.get('asideLabel', 'Aside')
                    elif issue_id in ['DiscoSectionFound', 'landmarks_DiscoSectionFound']:
                        signature = metadata.get('sectionSignature')
                        component_type = 'Section'
                        label = metadata.get('sectionLabel', 'Section')
                    elif issue_id in ['DiscoHeaderFound', 'landmarks_DiscoHeaderFound']:
                        signature = metadata.get('headerSignature')
                        component_type = 'Header'
                        label = metadata.get('headerLabel', 'Header')

                    if signature and signature != 'unknown':
                        if signature not in common_components:
                            common_components[signature] = {
                                'type': component_type,
                                'label': label,
                                'signature': signature,  # Store signature for display
                                'xpaths_by_page': {},  # page_url -> xpath
                                'pages': set()
                            }

                        xpath = d_dict.get('xpath', '') or metadata.get('xpath', '')
                        common_components[signature]['xpaths_by_page'][page_url] = xpath
                        common_components[signature]['pages'].add(page_url)

        return common_components

    def _create_project_deduped_issues_sheet(self, ws, data, styles):
        """Create a deduplicated issues sheet that groups issues by common components"""
        headers = ['Type', 'Impact', 'Rule ID', 'Touchpoint', 'What', 'Why Important', 'Who Affected',
                   'How to Remediate', 'WCAG Criteria', 'Location (XPath)', 'Element',
                   'Common Component(s)', 'Pages with Issue', 'Page Count', 'Breakpoints', 'Pseudoclasses', 'Page States', 'Test Users', 'User Roles']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, styles['header'])

        row = 2

        # Extract common components from discovery issues
        common_components = self._extract_common_components(data)

        # Track unique issues: (rule_id, xpath_or_component) -> issue data
        unique_issues = {}

        # Iterate through all websites and their pages
        for website_data in data.get('websites', []):
            website = website_data.get('website', {})
            website_name = website.get('name', '') if isinstance(website, dict) else getattr(website, 'name', '')

            for page_result in website_data.get('pages', []):
                page = page_result.get('page', {})
                page_url = page.get('url', '') if isinstance(page, dict) else getattr(page, 'url', '')

                test_result = page_result.get('test_result')
                if not test_result:
                    continue

                # Get page state description from test result if available
                page_state_desc = ''
                if test_result:
                    page_state = getattr(test_result, 'page_state', None) if hasattr(test_result, 'page_state') else test_result.get('page_state')
                    if page_state:
                        if isinstance(page_state, dict):
                            page_state_desc = page_state.get('description', '')
                        elif hasattr(page_state, 'description'):
                            page_state_desc = page_state.description

                # Process all issue types
                for issue_type, issue_list_attr in [('violation', 'violations'), ('warning', 'warnings'),
                                                      ('info', 'info'), ('discovery', 'discovery')]:
                    issues = getattr(test_result, issue_list_attr, []) if hasattr(test_result, issue_list_attr) else []

                    for issue in issues:
                        if hasattr(issue, 'to_dict'):
                            issue_dict = issue.to_dict()
                        else:
                            issue_dict = issue if isinstance(issue, dict) else {}

                        # Enrich with catalog information
                        issue_dict = IssueCatalog.enrich_issue(issue_dict)

                        rule_id = issue_dict.get('id', '')
                        issue_xpath = issue_dict.get('xpath', '')

                        # Get metadata (breakpoint, pseudoclass)
                        metadata = issue_dict.get('metadata', {})
                        breakpoint = metadata.get('breakpoint', '')
                        pseudoclass = metadata.get('pseudoclass', '')

                        # Find which common components contain this issue
                        containing_components = []
                        for signature, comp_data in common_components.items():
                            # Check if this issue is within this component on this page
                            comp_xpath = comp_data['xpaths_by_page'].get(page_url)
                            if comp_xpath and self._xpath_is_within(issue_xpath, comp_xpath):
                                # Format like Discovery Report: "Type signature"
                                containing_components.append(f"{comp_data['type']} {comp_data['signature']}")

                        # Create deduplication key
                        if containing_components:
                            # Dedupe by rule_id + component(s)
                            dedup_key = (rule_id, tuple(sorted(containing_components)))
                        else:
                            # Dedupe by rule_id + exact xpath for non-component issues
                            dedup_key = (rule_id, issue_xpath)

                        if dedup_key not in unique_issues:
                            unique_issues[dedup_key] = {
                                'type': issue_type,
                                'data': issue_dict,
                                'component': ', '.join(containing_components) if containing_components else '',
                                'pages': set(),
                                'page_xpaths': {},  # page -> xpath mapping
                                'breakpoints': set(),  # Track all breakpoints where this issue appears
                                'pseudoclasses': set(),  # Track all pseudoclasses
                                'page_states': set(),  # Track all page states
                                'test_users': set(),  # Track all test users who encountered this
                                'user_roles': set()  # Track all unique user roles
                            }

                        unique_issues[dedup_key]['pages'].add(page_url)
                        unique_issues[dedup_key]['page_xpaths'][page_url] = issue_xpath

                        # Add metadata to tracking sets
                        if breakpoint:
                            unique_issues[dedup_key]['breakpoints'].add(breakpoint)
                        if pseudoclass:
                            unique_issues[dedup_key]['pseudoclasses'].add(pseudoclass)
                        if page_state_desc:
                            unique_issues[dedup_key]['page_states'].add(page_state_desc)

                        # Track authenticated user info
                        auth_user = metadata.get('authenticated_user', {})
                        if auth_user:
                            unique_issues[dedup_key]['test_users'].add(auth_user.get('display_name', ''))
                            for role in auth_user.get('roles', []):
                                unique_issues[dedup_key]['user_roles'].add(role)
                        else:
                            unique_issues[dedup_key]['test_users'].add('Guest')

        # Write deduplicated issues to sheet
        for (rule_id, dedup_value), issue_data in sorted(unique_issues.items(),
                                                          key=lambda x: (x[1]['type'], x[0][0])):
            v_dict = issue_data['data']
            issue_type = issue_data['type']

            # Determine fill color based on type
            if issue_type == 'violation':
                fill_color = styles['violation']['fill']
                type_label = 'Violation'
            elif issue_type == 'warning':
                fill_color = styles['warning']['fill']
                type_label = 'Warning'
            elif issue_type == 'info':
                fill_color = styles['info']['fill']
                type_label = 'Info'
            else:  # discovery
                fill_color = styles.get('discovery', {}).get('fill', styles['info']['fill'])
                type_label = 'Discovery'

            ws.cell(row=row, column=1, value=type_label)
            ws.cell(row=row, column=2, value=str(v_dict.get('impact', 'Unknown')).upper())
            ws.cell(row=row, column=3, value=rule_id)
            ws.cell(row=row, column=4, value=v_dict.get('touchpoint', v_dict.get('category', '')))
            ws.cell(row=row, column=5, value=v_dict.get('description_full', v_dict.get('what', v_dict.get('description', ''))))
            ws.cell(row=row, column=6, value=v_dict.get('why_it_matters', ''))
            ws.cell(row=row, column=7, value=v_dict.get('who_it_affects', ''))
            ws.cell(row=row, column=8, value=v_dict.get('how_to_fix', v_dict.get('remediation', v_dict.get('suggested_fix', ''))))
            ws.cell(row=row, column=9, value=v_dict.get('wcag_full', ', '.join(v_dict.get('wcag_criteria', [])) if isinstance(v_dict.get('wcag_criteria'), list) else v_dict.get('wcag_criteria', '')))

            # For location, show one representative xpath
            representative_xpath = list(issue_data['page_xpaths'].values())[0] if issue_data['page_xpaths'] else ''
            ws.cell(row=row, column=10, value=representative_xpath)

            ws.cell(row=row, column=11, value=v_dict.get('element', ''))
            ws.cell(row=row, column=12, value=issue_data['component'])

            # List all pages where this issue appears
            pages_list = '\n'.join(sorted(issue_data['pages']))
            ws.cell(row=row, column=13, value=pages_list)
            ws.cell(row=row, column=13).alignment = self.Alignment(wrap_text=True, vertical='top')

            ws.cell(row=row, column=14, value=len(issue_data['pages']))

            # Add metadata columns - show all unique values across all instances of this issue
            breakpoints_list = ', '.join(sorted([str(bp) for bp in issue_data['breakpoints']])) if issue_data['breakpoints'] else ''
            ws.cell(row=row, column=15, value=breakpoints_list)

            pseudoclasses_list = ', '.join(sorted([str(pc) for pc in issue_data['pseudoclasses']])) if issue_data['pseudoclasses'] else ''
            ws.cell(row=row, column=16, value=pseudoclasses_list)

            page_states_list = ', '.join(sorted([str(ps) for ps in issue_data['page_states']])) if issue_data['page_states'] else ''
            ws.cell(row=row, column=17, value=page_states_list)
            ws.cell(row=row, column=17).alignment = self.Alignment(wrap_text=True, vertical='top')

            # Add test user columns
            test_users_list = ', '.join(sorted(issue_data['test_users'])) if issue_data['test_users'] else ''
            ws.cell(row=row, column=18, value=test_users_list)

            user_roles_list = ', '.join(sorted(issue_data['user_roles'])) if issue_data['user_roles'] else ''
            ws.cell(row=row, column=19, value=user_roles_list)

            # Apply fill color
            for col in range(1, 20):
                ws.cell(row=row, column=col).fill = fill_color

            row += 1

        self._auto_adjust_columns(ws)

    def _create_website_summary_sheet(self, ws, data, styles):
        """Create website summary sheet"""
        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f"Website Report - {data.get('website', {}).get('name', 'Unknown')}"
        title_cell.font = self.Font(bold=True, size=14)
        title_cell.alignment = self.Alignment(horizontal="center")
        
        # Statistics
        row = 3
        stats = data.get('statistics', {})
        
        ws.cell(row=row, column=1, value="Total Pages:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=stats.get('total_pages', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Total Violations:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=stats.get('total_violations', 0))
        ws.cell(row=row, column=2).fill = styles['violation']['fill']
        row += 1
        
        ws.cell(row=row, column=1, value="Total Warnings:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=stats.get('total_warnings', 0))
        ws.cell(row=row, column=2).fill = styles['warning']['fill']
        row += 1
        
        ws.cell(row=row, column=1, value="Average Violations/Page:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=f"{stats.get('average_violations', 0):.1f}")
        
        self._auto_adjust_columns(ws)
    
    def _create_pages_sheet(self, ws, pages, styles):
        """Create pages sheet"""
        headers = ['Page URL', 'Violations', 'Warnings', 'Passes', 'Last Tested', 'Page State', 'State Sequence', 'Session ID']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, styles['header'])

        row = 2
        for pr in pages:
            page = pr.get('page', {})
            test = pr.get('test_result', {})

            ws.cell(row=row, column=1, value=page.get('url', ''))

            violations_cell = ws.cell(row=row, column=2, value=test.get('violation_count', 0) if isinstance(test, dict) else getattr(test, 'violation_count', 0))
            if (test.get('violation_count', 0) if isinstance(test, dict) else getattr(test, 'violation_count', 0)) > 0:
                violations_cell.fill = styles['violation']['fill']

            warnings_cell = ws.cell(row=row, column=3, value=test.get('warning_count', 0) if isinstance(test, dict) else getattr(test, 'warning_count', 0))
            if (test.get('warning_count', 0) if isinstance(test, dict) else getattr(test, 'warning_count', 0)) > 0:
                warnings_cell.fill = styles['warning']['fill']

            ws.cell(row=row, column=4, value=test.get('pass_count', 0) if isinstance(test, dict) else getattr(test, 'pass_count', 0))
            ws.cell(row=row, column=5, value=str(test.get('test_date', '') if isinstance(test, dict) else getattr(test, 'test_date', '')))

            # Add multi-state information
            page_state = test.get('page_state') if isinstance(test, dict) else getattr(test, 'page_state', None)
            if page_state:
                state_desc = page_state.get('description', '') if isinstance(page_state, dict) else getattr(page_state, 'description', '')
                ws.cell(row=row, column=6, value=state_desc)
            else:
                ws.cell(row=row, column=6, value='')

            state_seq = test.get('state_sequence', '') if isinstance(test, dict) else getattr(test, 'state_sequence', '')
            ws.cell(row=row, column=7, value=state_seq if state_seq != 0 else '')

            session_id = test.get('session_id', '') if isinstance(test, dict) else getattr(test, 'session_id', '')
            ws.cell(row=row, column=8, value=session_id or '')

            row += 1

        self._auto_adjust_columns(ws)
    
    def _create_violation_types_sheet(self, ws, violation_types, styles):
        """Create violation types sheet"""
        headers = ['Rule ID', 'Count', 'Description', 'Pages Affected']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, styles['header'])
        
        row = 2
        sorted_types = sorted(violation_types.items(), key=lambda x: x[1]['count'], reverse=True)
        
        for rule_id, info in sorted_types:
            ws.cell(row=row, column=1, value=rule_id)
            ws.cell(row=row, column=2, value=info['count'])
            ws.cell(row=row, column=3, value=info.get('description', ''))
            ws.cell(row=row, column=4, value=len(set(info.get('pages', []))))
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_project_summary_sheet(self, ws, data, styles):
        """Create project summary sheet"""
        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f"Project Report - {data.get('project', {}).get('name', 'Unknown')}"
        title_cell.font = self.Font(bold=True, size=14)
        title_cell.alignment = self.Alignment(horizontal="center")
        
        # Statistics
        row = 3
        stats = data.get('statistics', {})
        
        ws.cell(row=row, column=1, value="Total Websites:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=stats.get('total_websites', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Total Pages:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=stats.get('total_pages', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Total Violations:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=stats.get('total_violations', 0))
        ws.cell(row=row, column=2).fill = styles['violation']['fill']
        row += 1
        
        ws.cell(row=row, column=1, value="Average Violations/Page:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=f"{stats.get('average_violations_per_page', 0):.1f}")
        
        self._auto_adjust_columns(ws)
    
    def _create_websites_sheet(self, ws, websites, styles):
        """Create websites sheet"""
        headers = ['Website Name', 'URL', 'Pages', 'Total Violations', 'Total Warnings']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, styles['header'])
        
        row = 2
        for wd in websites:
            website = wd.get('website', {})
            pages = wd.get('pages', [])
            
            # Handle test_result as object or dict
            total_violations = 0
            total_warnings = 0
            for p in pages:
                test_result = p.get('test_result')
                if test_result:
                    if hasattr(test_result, 'violation_count'):
                        # It's an object
                        total_violations += test_result.violation_count
                        total_warnings += test_result.warning_count
                    elif isinstance(test_result, dict):
                        # It's a dictionary
                        total_violations += test_result.get('violation_count', 0)
                        total_warnings += test_result.get('warning_count', 0)
            
            # Handle website as object or dict
            if hasattr(website, 'name'):
                # It's an object
                ws.cell(row=row, column=1, value=website.name or '')
                ws.cell(row=row, column=2, value=getattr(website, 'url', None) or getattr(website, 'base_url', ''))
            else:
                # It's a dictionary
                ws.cell(row=row, column=1, value=website.get('name', ''))
                ws.cell(row=row, column=2, value=website.get('url', website.get('base_url', '')))
            ws.cell(row=row, column=3, value=len(pages))
            
            violations_cell = ws.cell(row=row, column=4, value=total_violations)
            if total_violations > 0:
                violations_cell.fill = styles['violation']['fill']
            
            warnings_cell = ws.cell(row=row, column=5, value=total_warnings)
            if total_warnings > 0:
                warnings_cell.fill = styles['warning']['fill']
            
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_all_projects_summary_sheet(self, ws, data, styles):
        """Create all projects summary sheet"""
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "All Projects Accessibility Report"
        title_cell.font = self.Font(bold=True, size=14)
        title_cell.alignment = self.Alignment(horizontal="center")
        
        # Overall Statistics
        row = 3
        summary = data.get('summary', {})
        
        ws.cell(row=row, column=1, value="Overall Statistics").font = self.Font(bold=True, size=12)
        row += 1
        
        ws.cell(row=row, column=1, value="Total Projects:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('total_projects', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Total Websites:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('total_websites', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Total Pages:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('total_pages', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Pages Tested:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('total_tested', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Total Violations:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('total_violations', 0))
        ws.cell(row=row, column=2).fill = styles['violation']['fill']
        row += 1
        
        ws.cell(row=row, column=1, value="Total Warnings:").font = self.Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('total_warnings', 0))
        ws.cell(row=row, column=2).fill = styles['warning']['fill']
        
        self._auto_adjust_columns(ws)
    
    def _create_projects_breakdown_sheet(self, ws, projects, styles):
        """Create projects breakdown sheet"""
        headers = ['Project Name', 'Description', 'Websites', 'Total Pages', 'Tested Pages', 'Coverage %', 'Violations', 'Warnings']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, styles['header'])
        
        row = 2
        for project_data in projects:
            project = project_data.get('project', {})
            stats = project_data.get('stats', {})
            
            ws.cell(row=row, column=1, value=project.get('name', ''))
            ws.cell(row=row, column=2, value=project.get('description', ''))
            ws.cell(row=row, column=3, value=stats.get('website_count', 0))
            ws.cell(row=row, column=4, value=stats.get('total_pages', 0))
            ws.cell(row=row, column=5, value=stats.get('tested_pages', 0))
            ws.cell(row=row, column=6, value=f"{stats.get('test_coverage', 0):.1f}%")
            
            violations = sum(w.get('violations', 0) for w in project_data.get('websites', []))
            violations_cell = ws.cell(row=row, column=7, value=violations)
            if violations > 0:
                violations_cell.fill = styles['violation']['fill']
            
            warnings = sum(w.get('warnings', 0) for w in project_data.get('websites', []))
            warnings_cell = ws.cell(row=row, column=8, value=warnings)
            if warnings > 0:
                warnings_cell.fill = styles['warning']['fill']
            
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _apply_style(self, cell, style):
        """Apply style dictionary to a cell"""
        for attr, value in style.items():
            setattr(cell, attr, value)
    
    def _create_page_states_sheet(self, ws, data, styles):
        """Create sheet showing multi-state test results summary"""
        # Title
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = "Page State Information"
        title_cell.font = self.Font(bold=True, size=14)
        title_cell.alignment = self.Alignment(horizontal="center")

        # Headers
        row = 3
        headers = ['State Sequence', 'State Description', 'Errors', 'Warnings', 'Info', 'Discovery', 'Test Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, styles['header'])

        row += 1

        # Get test result
        test_result = data.get('test_result', {})
        session_id = test_result.get('session_id')

        # Show the current state
        state_sequence = test_result.get('state_sequence', 0)
        page_state = test_result.get('page_state', {})

        if isinstance(page_state, dict):
            state_desc = page_state.get('description', f'State {state_sequence}')
            scripts_executed = page_state.get('scripts_executed', [])
            elements_clicked = page_state.get('elements_clicked', [])
        else:
            state_desc = f'State {state_sequence}'
            scripts_executed = []
            elements_clicked = []

        ws.cell(row=row, column=1, value=state_sequence)
        ws.cell(row=row, column=2, value=state_desc)
        ws.cell(row=row, column=3, value=len(data.get('violations', [])))
        ws.cell(row=row, column=4, value=len(data.get('warnings', [])))
        ws.cell(row=row, column=5, value=len(data.get('info', [])))
        ws.cell(row=row, column=6, value=len(data.get('discovery', [])))
        ws.cell(row=row, column=7, value=str(test_result.get('test_date', '')))

        # Add state details section
        row += 2
        ws.cell(row=row, column=1, value="State Details:").font = self.Font(bold=True, size=12)
        row += 1

        if session_id:
            ws.cell(row=row, column=1, value="Session ID:").font = self.Font(bold=True)
            ws.cell(row=row, column=2, value=session_id)
            row += 1

        if scripts_executed:
            ws.cell(row=row, column=1, value="Scripts Executed:").font = self.Font(bold=True)
            ws.cell(row=row, column=2, value=', '.join(scripts_executed) if isinstance(scripts_executed, list) else str(scripts_executed))
            row += 1

        if elements_clicked:
            ws.cell(row=row, column=1, value="Elements Clicked:").font = self.Font(bold=True)
            if isinstance(elements_clicked, list) and len(elements_clicked) > 0:
                click_desc = ', '.join([str(el.get('description', el.get('selector', str(el)))) if isinstance(el, dict) else str(el) for el in elements_clicked])
                ws.cell(row=row, column=2, value=click_desc)
            row += 1

        # Add note about multi-state testing
        row += 1
        ws.cell(row=row, column=1, value="Note:").font = self.Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Multi-State Testing:")
        ws.cell(row=row, column=2, value="This report shows results for a single page state. If multiple states were tested during the same session, each state will have its own test result with a different state_sequence value.")
        row += 1
        ws.cell(row=row, column=1, value="Breakpoint Testing:")
        ws.cell(row=row, column=2, value="Some issues may have been detected at specific breakpoints (viewport widths). Check the 'Breakpoint (px)' column in the 'All Issues' sheet for breakpoint-specific issues.")
        row += 1
        ws.cell(row=row, column=1, value="Context Information:")
        ws.cell(row=row, column=2, value="Check the 'All Issues' sheet for complete context including Page State, Breakpoint, and Pseudoclass information for each issue.")

        self._auto_adjust_columns(ws)

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = self.get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


class PDFFormatter(BaseFormatter):
    """PDF report formatter (uses HTML + conversion)"""
    
    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.extension = 'pdf'
        self.html_formatter = HTMLFormatter(config)
        
        # Try to import weasyprint
        try:
            from weasyprint import HTML, CSS
            self.HTML = HTML
            self.CSS = CSS
            self.has_weasyprint = True
        except ImportError:
            logger.warning("weasyprint not installed - PDF generation will fall back to HTML")
            self.has_weasyprint = False
    
    def format_page_report(self, data: Dict[str, Any]) -> bytes:
        """Generate PDF for page report"""
        html_content = self.html_formatter.format_page_report(data)
        return self._convert_to_pdf(html_content)
    
    def format_website_report(self, data: Dict[str, Any]) -> bytes:
        """Generate PDF for website report"""
        html_content = self.html_formatter.format_website_report(data)
        return self._convert_to_pdf(html_content)
    
    def format_project_report(self, data: Dict[str, Any]) -> bytes:
        """Generate PDF for project report"""
        html_content = self.html_formatter.format_project_report(data)
        return self._convert_to_pdf(html_content)
    
    def format_all_projects_report(self, data: Dict[str, Any]) -> bytes:
        """Generate PDF for all projects report"""
        html_content = self.html_formatter.format_all_projects_report(data)
        return self._convert_to_pdf(html_content)
    
    def format_summary_report(self, data: Dict[str, Any]) -> bytes:
        """Generate PDF for summary report"""
        html_content = self.html_formatter.format_summary_report(data)
        return self._convert_to_pdf(html_content)
    
    def _convert_to_pdf(self, html_content: str) -> bytes:
        """Convert HTML content to PDF bytes"""
        if self.has_weasyprint:
            try:
                # Add some PDF-specific CSS for better rendering
                pdf_css = self.CSS(string='''
                    @page {
                        size: A4;
                        margin: 1cm;
                    }
                    body {
                        font-size: 10pt;
                    }
                    .container {
                        max-width: 100%;
                        box-shadow: none;
                    }
                    table {
                        page-break-inside: avoid;
                    }
                    .violation, .warning, .pass, .ai-finding {
                        page-break-inside: avoid;
                    }
                ''')
                
                # Create PDF from HTML
                pdf_document = self.HTML(string=html_content).render(stylesheets=[pdf_css])
                pdf_bytes = pdf_document.write_pdf()
                
                return pdf_bytes
            except Exception as e:
                logger.error(f"Failed to generate PDF with weasyprint: {e}")
                # Fall back to returning HTML as bytes
                return html_content.encode('utf-8')
        else:
            # If weasyprint is not available, return HTML as bytes
            logger.warning("PDF generation not available - returning HTML content")
            return html_content.encode('utf-8')
    
    def save_pdf(self, html_content: str, filepath: Path):
        """
        Save HTML as PDF
        
        This method is deprecated - use the format methods that return bytes instead
        """
        pdf_bytes = self._convert_to_pdf(html_content)
        with open(filepath, 'wb') as f:
            f.write(pdf_bytes)